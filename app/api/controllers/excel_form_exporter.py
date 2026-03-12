"""
FormExcelExporter v6.0
Replica la lógica de renderizado de FormResponseRenderer (ResponsesModal.tsx).
Genera un xlsx donde cada fila de respuesta queda claramente estructurada,
los repeaters se muestran como sub-tablas y el header/footer se replica.
"""
from io import BytesIO
from typing import Any, Dict, List, Optional
import json

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None  # type: ignore


# ── colores (mismo que frontend) ─────────────────────────────────────────────
C_TEAL_DARK  = "0f8594"   # repeater header bg
C_TEAL_LIGHT = "f0fdfa"   # sub-repeater header bg
C_TEAL_TEXT  = "0f766e"   # sub-repeater header text
C_GRAY_HEADER= "f3f4f6"   # column headers bg
C_GRAY_TEXT  = "374151"   # column headers text
C_ALT_ROW    = "f9fafb"   # alternating row
C_BORDER     = "d1d5db"   # default border
C_EMPTY      = "9ca3af"   # "sin respuesta"
C_WHITE      = "ffffff"
C_LABEL_BG   = "f9fafb"
C_LABEL_BORDER = "e5e7eb"


def _thin(color: str = C_BORDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, italic: bool = False, color: str = "000000",
          size: int = 10) -> Font:
    return Font(bold=bold, italic=italic, color=color, size=size, name="Calibri")


def _align(h: str = "left", v: str = "center", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── formateadores (misma lógica que PDF) ─────────────────────────────────────

def _fmt_firm_text(answer_text: str) -> str:
    """Devuelve texto plano de firma para celda Excel."""
    try:
        data = json.loads(answer_text)
        firm = data.get("firmData", {})
        if isinstance(firm, dict) and firm.get("qr_url"):
            name = firm.get("person_name", "")
            pid  = firm.get("person_id", "")
        elif isinstance(firm, dict) and isinstance(firm.get("firmData"), dict):
            inner = firm["firmData"]
            name  = inner.get("person_name", "")
            pid   = inner.get("person_id", "")
            firm  = inner
        else:
            name = pid = ""
        qr = firm.get("qr_url", "")
        parts = ["✓ Firma validada"]
        if name:
            parts.append("por " + name)
        if pid:
            parts.append("ID: " + str(pid))
        if qr:
            parts.append("QR: " + qr)
        return " | ".join(parts)
    except Exception:
        return str(answer_text)


def _fmt_location_text(answer_text: str) -> str:
    try:
        loc = json.loads(answer_text)
        parts = []
        if loc.get("selection"):
            parts.append(str(loc["selection"]))
        coords = []
        if loc.get("lat") is not None:
            coords.append("Lat: " + str(loc["lat"]))
        if loc.get("lng") is not None:
            coords.append("Lng: " + str(loc["lng"]))
        if coords:
            parts.append(" | ".join(coords))
        if loc.get("address"):
            parts.append(str(loc["address"]))
        return " — ".join(parts) if parts else str(answer_text)
    except Exception:
        return str(answer_text)


def _fmt_checkbox_text(answer_text: str) -> str:
    try:
        vals = json.loads(answer_text)
        if isinstance(vals, list):
            return ", ".join(str(v) for v in vals)
    except Exception:
        pass
    return str(answer_text)


def _fmt_number_text(answer_text: str, props: dict) -> str:
    try:
        num      = float(answer_text)
        prefix   = str(props.get("currencyPrefix") or "")
        suffix   = str(props.get("currencySuffix") or "")
        decimals = int(props.get("decimalPlaces") or 2)
        return prefix + "{:,.{d}f}".format(num, d=decimals) + suffix
    except Exception:
        return str(answer_text)


def _fmt_date_text(answer_text: str) -> str:
    from datetime import datetime
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M"):
        try:
            return datetime.strptime(answer_text[:19], fmt).strftime("%d/%m/%Y")
        except Exception:
            pass
    return str(answer_text)


def _fmt_datetime_text(answer_text: str) -> str:
    from datetime import datetime
    try:
        s = answer_text.replace("T", " ").split(".")[0]
        return datetime.strptime(s[:16], "%Y-%m-%d %H:%M").strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(answer_text)


def _cell_value_text(cell_data: Any) -> str:
    """Texto plano para celdas de repeater (renderRepeaterCell)."""
    if cell_data is None:
        return "-"
    if isinstance(cell_data, str):
        try:
            p = json.loads(cell_data)
            fd = p.get("firmData", {})
            if isinstance(fd, dict) and (fd.get("qr_url") or
                    (isinstance(fd.get("firmData"), dict) and fd["firmData"].get("qr_url"))):
                return _fmt_firm_text(cell_data)
        except Exception:
            pass
        return cell_data or "-"
    if isinstance(cell_data, dict):
        qtype = cell_data.get("question_type", "")
        atext = str(cell_data.get("answer_text") or "")
        fpath = str(cell_data.get("file_path") or "")
        if qtype == "firm" and atext:
            try:
                return _fmt_firm_text(atext)
            except Exception:
                pass
        if atext:
            try:
                p = json.loads(atext)
                fd = p.get("firmData", {})
                if isinstance(fd, dict) and (fd.get("qr_url") or
                        (isinstance(fd.get("firmData"), dict) and fd["firmData"].get("qr_url"))):
                    return _fmt_firm_text(atext)
            except Exception:
                pass
        parts = []
        if atext:
            parts.append(atext)
        if fpath:
            parts.append("[Archivo: " + fpath + "]")
        return " | ".join(parts) if parts else "-"
    return str(cell_data)


def _build_sub_rows_excel(filtered: list, sub_normal: list) -> list:
    """
    Puerto EXACTO de subRows en renderSubRepeaterForRow (ResponsesModal.tsx línea 958-1000).
    Sin repeater_row_index → reconstruir por columna+posición (igual que frontend).
    """
    sub_rows: list = []
    has_ri = any(a.get("repeater_row_index") is not None for a in filtered)

    if has_ri:
        by_ri: dict = {}
        for ans in filtered:
            by_ri.setdefault(ans.get("repeater_row_index") or 0, []).append(ans)
        for ri in sorted(by_ri):
            rd: dict = {}
            for c in sub_normal:
                cid = c.get("id", "")
                m   = next((a for a in by_ri[ri] if a.get("form_design_element_id") == cid), None)
                if m:
                    rd[cid] = {"answer_text": m.get("answer_text"),
                               "file_path":   m.get("file_path"),
                               "question_type": m.get("question_type")}
            if rd:
                sub_rows.append(rd)
    else:
        # Sin repeater_row_index: agrupar por COLUMNA y reconstruir por posición (frontend línea 979-998)
        by_col: dict = {}
        for c in sub_normal:
            cid = c.get("id", "")
            by_col[cid] = sorted(
                [a for a in filtered if a.get("form_design_element_id") == cid],
                key=lambda a: (a.get("id_answer") or 0),
            )
        max_r = max((len(v) for v in by_col.values()), default=0)
        for i in range(max_r):
            rd = {}
            for c in sub_normal:
                cid = c.get("id", "")
                if i < len(by_col.get(cid, [])):
                    a = by_col[cid][i]
                    rd[cid] = {"answer_text": a.get("answer_text"),
                               "file_path":   a.get("file_path"),
                               "question_type": a.get("question_type")}
            if rd:
                sub_rows.append(rd)

    return sub_rows


# ── Clase principal ───────────────────────────────────────────────────────────

class FormExcelExporter:
    """
    Genera Excel de una respuesta de formulario
    replicando la lógica de FormResponseRenderer.
    """

    def __init__(
        self,
        form_design: list,
        answers: list,
        style_config: Optional[dict] = None,
        form_title: str = "",
        submitted_at: str = "",
        response_id: Optional[int] = None,
    ):
        if openpyxl is None:
            raise RuntimeError("openpyxl no instalado: pip install openpyxl")

        self.form_design  = form_design
        self.answers      = answers
        self.style_config = style_config or {}
        self.form_title   = form_title
        self.submitted_at = submitted_at
        self.response_id  = response_id

        # answersMap — misma prioridad que el frontend
        self._answers_map: Dict[str, Any] = {}
        for ans in answers:
            if ans.get("question_text"):
                self._answers_map[str(ans["question_text"])] = ans
            if ans.get("question_id") is not None:
                self._answers_map[str(ans["question_id"])] = ans
            if ans.get("form_design_element_id"):
                self._answers_map[str(ans["form_design_element_id"])] = ans

        self._wb  = openpyxl.Workbook()
        self._ws  = self._wb.active
        self._ws.title = "Respuesta"
        self._row = 1           # fila actual en la hoja
        self._col_width = 30.0  # ancho default por columna

    # ── getAnswerForField ─────────────────────────────────────────────────────

    def _get_answer(self, field: dict) -> Optional[dict]:
        fid   = str(field.get("id") or "")
        props = field.get("props") or {}
        if fid and fid in self._answers_map:
            return self._answers_map[fid]
        lex = str(field.get("linkExternalId") or "")
        if lex and lex in self._answers_map:
            return self._answers_map[lex]
        sqid = str(props.get("sourceQuestionId") or "")
        if sqid and sqid in self._answers_map:
            return self._answers_map[sqid]
        lbl = str(props.get("label") or "")
        if lbl and lbl in self._answers_map:
            return self._answers_map[lbl]
        return None

    # ── shouldRenderLayout ────────────────────────────────────────────────────

    def _find_recursive(self, items: list, qid: str) -> Optional[dict]:
        for item in items:
            if str(item.get("id_question", "")) == qid:
                return item
            found = self._find_recursive(item.get("children") or [], qid)
            if found:
                return found
        return None

    def _should_render(self, layout: dict) -> bool:
        props = layout.get("props") or {}
        if not props.get("hidden"):
            return True
        condicion = str(props.get("condicion") or "")
        valor     = str(props.get("valor") or "")
        if not condicion or not valor:
            return False
        parts = condicion.split("-")
        if len(parts) < 2:
            return False
        cond_field = self._find_recursive(self.form_design, parts[1])
        if not cond_field:
            return False
        answer  = self._get_answer(cond_field)
        current = str((answer or {}).get("answer_text", "") or "")
        allowed = [v.strip() for v in valor.split(",")]
        return current in allowed

    # ── helpers de escritura ──────────────────────────────────────────────────

    def _write_cell(
        self, row: int, col: int, value: str,
        bold: bool = False, italic: bool = False,
        font_color: str = "000000", font_size: int = 10,
        bg_color: Optional[str] = None,
        h_align: str = "left", border: bool = True,
        wrap: bool = True, col_span: int = 1,
    ) -> None:
        ws = self._ws
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = _font(bold=bold, italic=italic, color=font_color, size=font_size)
        cell.alignment = _align(h=h_align, wrap=wrap)
        if bg_color:
            cell.fill = _fill(bg_color)
        if border:
            cell.border = _thin()
        if col_span > 1:
            end_col = col + col_span - 1
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=end_col,
            )
            for c in range(col + 1, end_col + 1):
                ws.cell(row=row, column=c).border = _thin()

    def _set_row_height(self, row: int, height: float = 20.0) -> None:
        self._ws.row_dimensions[row].height = height

    def _set_col_width(self, col: int, width: float) -> None:
        letter = get_column_letter(col)
        if self._ws.column_dimensions[letter].width < width:
            self._ws.column_dimensions[letter].width = width

    # ── Header (HeaderTable + logo) ───────────────────────────────────────────

    def _write_header(self) -> None:
        sc = self.style_config
        ht = sc.get("headerTable") or {}
        if ht.get("enabled") and ht.get("cells"):
            self._write_header_table(ht)
        elif self.form_title:
            self._write_cell(
                self._row, 1, self.form_title,
                bold=True, font_size=14,
                bg_color="0f8594", font_color=C_WHITE,
                h_align="center", col_span=6,
            )
            self._set_row_height(self._row, 28)
            self._row += 1

        if self.submitted_at or self.response_id:
            parts = []
            if self.response_id:
                parts.append("Respuesta #" + str(self.response_id))
            if self.submitted_at:
                parts.append("Enviado: " + str(self.submitted_at))
            self._write_cell(
                self._row, 1, "  ".join(parts),
                font_color="6B7280", italic=True,
                bg_color=C_LABEL_BG, col_span=6, border=False,
            )
            self._set_row_height(self._row, 16)
            self._row += 1

        self._row += 1  # espacio

    def _write_header_table(self, ht: dict) -> None:
        cells = ht.get("cells") or []
        for row_cells in cells:
            col = 1
            for cell in row_cells:
                cont  = cell.get("content", "")
                cs    = int(cell.get("colSpan", 1))
                cbg   = (cell.get("backgroundColor") or "f3f4f6").lstrip("#")
                cc    = (cell.get("color") or "374151").lstrip("#")
                fw    = cell.get("fontWeight", "normal") == "bold"
                is_logo = cell.get("customClass") == "logo-cell" or cont == "[LOGO]"
                if is_logo:
                    cont = "[Logo]"
                self._write_cell(
                    self._row, col, cont if not is_logo else "[Logo]",
                    bold=fw, font_color=cc, bg_color=cbg,
                    col_span=cs,
                )
                self._set_col_width(col, max(self._col_width, len(cont) + 4))
                col += cs
            self._set_row_height(self._row, 22)
            self._row += 1

    # ── Footer ────────────────────────────────────────────────────────────────

    def _write_footer(self) -> None:
        footer = (self.style_config or {}).get("footer") or {}
        if not footer.get("show"):
            return
        self._row += 1
        self._write_cell(
            self._row, 1, str(footer.get("text", "")),
            italic=True, font_color="9CA3AF",
            bg_color=C_LABEL_BG, col_span=6, border=False,
            h_align=footer.get("align", "center"),
        )
        self._set_row_height(self._row, 16)
        self._row += 1

    # ── renderDecorativeElement ───────────────────────────────────────────────

    def _write_decorative(self, field: dict) -> None:
        ftype = field.get("type", "")
        props = field.get("props") or {}

        if ftype == "label":
            text  = str(props.get("text") or "Etiqueta")
            bold  = str(props.get("fontWeight", "normal")).lower() in ("bold", "700", "600")
            color = str(props.get("color") or "333333").lstrip("#")
            self._write_cell(
                self._row, 1, text,
                bold=bold, font_color=color, font_size=12,
                bg_color=None, border=False, col_span=6,
            )
            self._set_row_height(self._row, 18)
            self._row += 1
            return

        if ftype == "helpText":
            text = str(props.get("text") or "")
            self._write_cell(
                self._row, 1, text,
                italic=True, font_color=C_EMPTY, font_size=9,
                bg_color=None, border=False, col_span=6,
            )
            self._set_row_height(self._row, 14)
            self._row += 1
            return

        if ftype == "divider":
            # fila vacía como separador visual
            self._set_row_height(self._row, 6)
            self._row += 1
            return

        if ftype == "image":
            lbl = str(props.get("label") or "Imagen")
            src = str(props.get("src") or "")
            self._write_cell(
                self._row, 1, "[Imagen: " + lbl + ("] → " + src if src else "]"),
                italic=True, font_color="6B7280",
                bg_color=C_LABEL_BG, border=False, col_span=6,
            )
            self._set_row_height(self._row, 16)
            self._row += 1
            return

        if ftype == "button":
            text = str(props.get("text") or "Botón")
            self._write_cell(
                self._row, 1, text + " (deshabilitado)",
                italic=True, font_color=C_EMPTY,
                bg_color=C_LABEL_BG, border=False, col_span=6,
            )
            self._set_row_height(self._row, 16)
            self._row += 1
            return

    # ── renderInputFieldResponse ──────────────────────────────────────────────

    def _write_input_field(self, field: dict) -> None:
        props  = field.get("props") or {}
        label  = str(props.get("label") or "Campo")
        req    = " *" if props.get("required") else ""
        answer = self._get_answer(field)
        ftype  = field.get("type", "")

        # columna A = etiqueta, columna B = respuesta
        self._write_cell(
            self._row, 1, label + req,
            bold=True, font_size=10,
            bg_color=C_LABEL_BG, col_span=1,
        )

        if answer is not None:
            atext = str(answer.get("answer_text") or "")
            fpath = str(answer.get("file_path") or "")
            qtype = str(answer.get("question_type") or ftype)

            if qtype == "firm" and atext:
                value = _fmt_firm_text(atext)
            elif qtype == "location" and atext:
                value = _fmt_location_text(atext)
            elif qtype in ("checkbox", "multiselect") and atext:
                value = _fmt_checkbox_text(atext)
            elif qtype == "number" and atext:
                value = _fmt_number_text(atext, props)
            elif qtype == "date" and atext:
                value = _fmt_date_text(atext)
            elif qtype in ("datetime", "datetimelocal") and atext:
                value = _fmt_datetime_text(atext)
            elif atext:
                value = atext
            else:
                value = ""

            if fpath:
                value = (value + " | " if value else "") + "[Archivo: " + fpath + "]"

            if not value:
                value = "Sin respuesta"

            self._write_cell(
                self._row, 2, value,
                font_color="1F2937", bg_color=C_WHITE, col_span=5,
            )
        else:
            self._write_cell(
                self._row, 2, "Sin respuesta registrada",
                italic=True, font_color=C_EMPTY,
                bg_color=C_WHITE, col_span=5,
            )

        self._set_row_height(self._row, 20)
        self._row += 1

    # ── renderRepeaterResponse ────────────────────────────────────────────────

    def _write_repeater(self, field: dict, indent_col: int = 1) -> None:
        props    = field.get("props") or {}
        children = field.get("children") or []

        normal_ch = [c for c in children if c.get("type") != "repeater"]
        sub_ch    = [c for c in children if c.get("type") == "repeater"]

        if not children:
            self._write_cell(
                self._row, indent_col, "Tabla sin columnas configuradas",
                italic=True, font_color=C_EMPTY, col_span=6,
            )
            self._row += 1
            return

        # Encabezado del repeater (teal)
        lbl = str(props.get("label") or "")
        if lbl:
            n_cols = max(len(normal_ch), 1)
            self._write_cell(
                self._row, indent_col, "▤ " + lbl.upper(),
                bold=True, font_color=C_WHITE,
                bg_color=C_TEAL_DARK, font_size=10,
                col_span=n_cols,
            )
            self._set_row_height(self._row, 20)
            self._row += 1

        # Construir parentRows — PORT EXACTO del frontend (ResponsesModal.tsx línea 735-808)
        col_ids  = [c["id"] for c in normal_ch if c.get("id")]
        col_qids = [
            str(c.get("linkExternalId") or (c.get("props") or {}).get("sourceQuestionId") or "")
            for c in normal_ch
        ]
        col_qids = [q for q in col_qids if q]

        rep_answers = [
            ans for ans in self.answers
            if (ans.get("repeated_id") not in (None, "", "null"))
            and not ans.get("parent_repeated_id")
            and (
                (ans.get("form_design_element_id") and ans["form_design_element_id"] in col_ids)
                or (ans.get("question_id") and str(ans["question_id"]) in col_qids)
            )
        ]

        # Ordenar antes de agrupar para estabilidad
        rep_answers = sorted(
            rep_answers,
            key=lambda a: (
                a.get("repeater_row_index") if a.get("repeater_row_index") is not None else 999999,
                a.get("id_answer") or 0,
            )
        )

        parent_rows: List[Dict] = []
        grouped: Dict[str, List] = {}
        group_first_ri: Dict[str, int] = {}
        for ans in rep_answers:
            rid = str(ans["repeated_id"])
            if rid not in grouped:
                grouped[rid] = []
                group_first_ri[rid] = (
                    int(ans["repeater_row_index"]) if ans.get("repeater_row_index") is not None
                    else (ans.get("id_answer") or 0)
                )
            grouped[rid].append(ans)

        for repeated_id, grp in grouped.items():
            by_col: Dict[str, List] = {}
            for child in normal_ch:
                cid  = child.get("id", "")
                qid  = str(child.get("linkExternalId") or (child.get("props") or {}).get("sourceQuestionId", "") or "")
                col  = [a for a in grp if a.get("form_design_element_id") == cid]
                if not col and qid:
                    col = [a for a in grp if str(a.get("question_id", "")) == qid]
                has_idx = any(a.get("repeater_row_index") is not None for a in col)
                col = sorted(col, key=lambda a: ((a.get("repeater_row_index") or 0) if has_idx else (a.get("id_answer") or 0)))
                by_col[cid] = col

            max_r = max((len(v) for v in by_col.values()), default=1)
            for row_index in range(max_r):
                row_data: Dict[str, Any] = {}
                for child in normal_ch:
                    cid = child.get("id", "")
                    col = by_col.get(cid, [])
                    if row_index < len(col):
                        a = col[row_index]
                        row_data[cid] = {
                            "answer_text":  a.get("answer_text"),
                            "file_path":    a.get("file_path"),
                            "question_type": a.get("question_type"),
                        }
                if row_data:
                    parent_rows.append({
                        "rowIndex":   row_index,   # índice LOCAL (igual que frontend)
                        "repeatedId": repeated_id,
                        "rowData":    row_data,
                    })

        # Sort por rowIndex, sort secundario por group_first_ri para estabilidad
        parent_rows.sort(key=lambda r: (r["rowIndex"], group_first_ri.get(r["repeatedId"], 999999)))

        n_parent_rows = len(parent_rows)

        if not normal_ch and not sub_ch:
            return

        # Encabezados de columnas
        if normal_ch:
            for ci, child in enumerate(normal_ch):
                c_lbl = str((child.get("props") or {}).get("label") or "Campo")
                c_req = " *" if (child.get("props") or {}).get("required") else ""
                self._write_cell(
                    self._row, indent_col + ci, c_lbl + c_req,
                    bold=True, font_color=C_GRAY_TEXT,
                    bg_color=C_GRAY_HEADER, font_size=10,
                )
                self._set_col_width(indent_col + ci, max(self._col_width, len(c_lbl) + 6))
            self._set_row_height(self._row, 20)
            self._row += 1

            if parent_rows:
                for disp_idx, prow in enumerate(parent_rows):
                    row_data    = prow["rowData"]
                    repeated_id = prow["repeatedId"]
                    row_idx     = prow["rowIndex"]    # índice LOCAL (igual que frontend)
                    bg = C_ALT_ROW if disp_idx % 2 == 1 else C_WHITE

                    for ci, child in enumerate(normal_ch):
                        cid      = child.get("id", "")
                        lex      = str(child.get("linkExternalId") or "")
                        cell_val = row_data.get(cid) or (row_data.get(lex) if lex else None)
                        text     = _cell_value_text(cell_val)
                        is_empty = (text == "-")
                        self._write_cell(
                            self._row, indent_col + ci, text,
                            italic=is_empty, font_color=C_EMPTY if is_empty else "1F2937",
                            bg_color=bg,
                        )
                        self._set_col_width(indent_col + ci, max(self._col_width, min(len(text) + 4, 60)))
                    self._set_row_height(self._row, 20)
                    self._row += 1

                    # sub-repeaters de esta fila
                    for sf in sub_ch:
                        self._write_sub_repeater(
                            sf, row_idx, repeated_id, indent_col + 1,
                            parent_field_id=str(field.get("id") or ""),
                            n_parent_rows=n_parent_rows,
                        )
            else:
                # sin datos
                self._write_cell(
                    self._row, indent_col, "Sin datos registrados",
                    italic=True, font_color=C_EMPTY,
                    bg_color=C_WHITE, col_span=len(normal_ch),
                )
                self._row += 1
        else:
            # Solo sub-repeaters sin columnas normales
            pfid = str(field.get("id") or "")
            if parent_rows:
                for prow in parent_rows:
                    self._write_cell(
                        self._row, indent_col,
                        "Registro " + str(prow["rowIndex"] + 1),
                        italic=True, font_color=C_EMPTY,
                        bg_color=C_LABEL_BG, border=False,
                    )
                    self._row += 1
                    for sf in sub_ch:
                        self._write_sub_repeater(
                            sf, prow["rowIndex"], prow["repeatedId"],
                            indent_col + 1,
                            parent_field_id=pfid,
                            n_parent_rows=n_parent_rows,
                        )
            else:
                for sf in sub_ch:
                    self._write_sub_repeater(sf, 0, "", indent_col + 1,
                                             parent_field_id=pfid, n_parent_rows=1)

        self._row += 1  # espacio post-repeater

    def _write_sub_repeater(
        self,
        sub_field: dict,
        parent_row_idx: int,           # rowIndex LOCAL del grupo
        parent_repeated_id: str,       # repeatedId del padre
        indent_col: int,
        parent_field_id: str = "",     # field.id del repeater PADRE (estrategia 2)
        n_parent_rows: int = 1,        # len(parentRows) — para estrategia 3
    ) -> None:
        sub_children = sub_field.get("children") or []
        sub_normal   = [c for c in sub_children if c.get("type") != "repeater"]
        sub_col_ids  = [c["id"] for c in sub_normal if c.get("id")]

        all_sub = [
            ans for ans in self.answers
            if ans.get("form_design_element_id") in sub_col_ids
            and ans.get("repeated_id") not in (None, "", "null")
        ]
        if not all_sub:
            return

        # ── ESTRATEGIA 1: parent_repeated_id === UUID de la fila padre ──────────
        filtered = [
            a for a in all_sub
            if str(a.get("parent_repeated_id") or "") == str(parent_repeated_id or "")
            and (parent_repeated_id not in (None, "", "null"))
        ]

        # ── ESTRATEGIA 2: parent_repeated_id === field.id + parent_row_index ────
        if not filtered and parent_field_id:
            filtered = [
                a for a in all_sub
                if str(a.get("parent_repeated_id") or "") == str(parent_field_id)
                and a.get("parent_row_index") == parent_row_idx
            ]

        # ── ESTRATEGIA 3: fallback exacto del frontend (línea 909-943) ──────────
        if not filtered and all_sub:
            col_count          = len(sub_col_ids) or 1
            total_sub_rows     = round(len(all_sub) / col_count)
            total_parent_rows  = max(1, n_parent_rows)
            sub_rows_per_parent = max(1, -(-total_sub_rows // total_parent_rows))
            sub_row_start      = parent_row_idx * sub_rows_per_parent
            sub_row_end        = sub_row_start + sub_rows_per_parent

            has_ri = any(a.get("repeater_row_index") is not None for a in all_sub)
            if has_ri:
                filtered = [
                    a for a in all_sub
                    if sub_row_start <= (a.get("repeater_row_index") or 0) < sub_row_end
                ]
            else:
                sorted_sub         = sorted(all_sub, key=lambda a: (a.get("id_answer") or 0))
                answers_per_parent = sub_rows_per_parent * col_count
                start_idx          = parent_row_idx * answers_per_parent
                filtered           = sorted_sub[start_idx: start_idx + answers_per_parent]

        if not filtered:
            return

        sub_rows = _build_sub_rows_excel(filtered, sub_normal)

        # Encabezado sub-repeater (teal light)
        lbl = str((sub_field.get("props") or {}).get("label") or "Sub-tabla")
        n_sub = max(len(sub_normal), 1)
        self._write_cell(
            self._row, indent_col, "↳ " + lbl.upper(),
            bold=True, font_color=C_TEAL_TEXT,
            bg_color=C_TEAL_LIGHT, font_size=9,
            col_span=n_sub,
        )
        self._set_row_height(self._row, 16)
        self._row += 1

        # Encabezados de columnas del sub
        for ci, child in enumerate(sub_normal):
            c_lbl = str((child.get("props") or {}).get("label") or "Campo")
            self._write_cell(
                self._row, indent_col + ci, c_lbl,
                bold=True, font_color=C_GRAY_TEXT,
                bg_color=C_GRAY_HEADER, font_size=9,
            )
            self._set_col_width(indent_col + ci, max(self._col_width, len(c_lbl) + 4))
        self._set_row_height(self._row, 16)
        self._row += 1

        for ri, rd in enumerate(sub_rows):
            bg = C_ALT_ROW if ri % 2 == 1 else C_WHITE
            for ci, child in enumerate(sub_normal):
                cid      = child.get("id", "")
                cell_val = rd.get(cid)
                text     = _cell_value_text(cell_val)
                is_empty = (text == "-")
                self._write_cell(
                    self._row, indent_col + ci, text,
                    italic=is_empty, font_color=C_EMPTY if is_empty else "1F2937",
                    bg_color=bg, font_size=9,
                )
            self._set_row_height(self._row, 16)
            self._row += 1

    # ── render field dispatch ─────────────────────────────────────────────────

    def _write_field(self, field: dict) -> None:
        ftype = field.get("type", "")
        if not ftype or not field.get("id"):
            return

        if ftype == "verticalLayout":
            if not self._should_render(field):
                return
            for child in (field.get("children") or []):
                self._write_field(child)
            return

        if ftype == "horizontalLayout":
            if not self._should_render(field):
                return
            for child in (field.get("children") or []):
                self._write_field(child)
            return

        if ftype in ("label", "helpText", "divider", "image", "button"):
            self._write_decorative(field)
            return

        if ftype == "repeater":
            self._write_repeater(field)
            return

        self._write_input_field(field)

    # ── renderFieldsWithAutoLayout ────────────────────────────────────────────

    def _write_all_fields(self) -> None:
        for field in self.form_design:
            self._write_field(field)

    # ── configurar anchos de columnas ─────────────────────────────────────────

    def _setup_columns(self) -> None:
        ws = self._ws
        # col 1: etiqueta
        ws.column_dimensions["A"].width = 28
        # cols 2-6: contenido
        for i in range(2, 8):
            letter = get_column_letter(i)
            if ws.column_dimensions[letter].width < 20:
                ws.column_dimensions[letter].width = 20

    # ── API pública ───────────────────────────────────────────────────────────

    def generate(self) -> BytesIO:
        self._write_header()
        self._write_all_fields()
        self._write_footer()
        self._setup_columns()

        # congelar primera fila si hay encabezado
        self._ws.freeze_panes = "A2"

        buf = BytesIO()
        self._wb.save(buf)
        buf.seek(0)
        return buf


def generate_form_excel(
    form_design: list,
    answers: list,
    style_config: Optional[dict] = None,
    form_title: str = "",
    submitted_at: str = "",
    response_id: Optional[int] = None,
) -> BytesIO:
    return FormExcelExporter(
        form_design=form_design,
        answers=answers,
        style_config=style_config,
        form_title=form_title,
        submitted_at=submitted_at,
        response_id=response_id,
    ).generate()