"""
Endpoint para exportar respuestas del usuario como plantilla Excel reimportable.
Se registra desde forms.py con register_export_template_route(router).
"""
from io import BytesIO
from typing import Optional

from fastapi import Query, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.core.security import get_current_user
from app.models import Answer, Form, Response, ResponseStatus, User


def register_export_template_route(router):
    """Call from forms.py to register the export-template GET endpoint."""

    @router.get("/{form_id}/responses/export-template-excel")
    def export_responses_as_template_excel(
        form_id: int,
        date_from: Optional[str] = Query(None, description="Fecha/hora inicio ISO (ej: 2025-01-15T08:00)"),
        date_to: Optional[str] = Query(None, description="Fecha/hora fin ISO (ej: 2025-01-31T23:59)"),
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
    ):
        """
        Genera un Excel con la misma estructura de la plantilla de importacion
        (filas 0-3 = ID Pregunta / Tipo / Nivel / Pregunta) pero relleno con las
        respuestas reales del usuario autenticado, filtradas opcionalmente por
        rango de fechas. El archivo resultante es reimportable con
        "Importar desde Excel".
        """
        import json
        from datetime import datetime as _dt
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        # ── 1. Formulario y form_design ───────────────────────────────────────
        form = db.query(Form).filter(Form.id == form_id).first()
        if not form:
            raise HTTPException(status_code=404, detail="Formulario no encontrado")

        fd = form.form_design
        if isinstance(fd, str):
            try:
                fd = json.loads(fd)
            except json.JSONDecodeError:
                fd = None
        if not fd or not isinstance(fd, list) or len(fd) == 0:
            raise HTTPException(
                status_code=400,
                detail="El formulario no tiene form_design",
            )

        # ── 2. Extraer campos ─────────────────────────────────────────────────
        # Replica la logica de buildFormTemplateWorkbook del frontend.
        # level: normal | repeater | sub-repeater
        # nivel: "" | repeaterId | repeaterId|subRepeaterId
        fields: list[dict] = []

        def _walk(item, rep_ctx=None, sub_ctx=None):
            t = item.get("type", "")
            if t in ("horizontalLayout", "verticalLayout"):
                for c in item.get("children") or []:
                    _walk(c, rep_ctx, sub_ctx)
            elif t == "repeater":
                if rep_ctx:
                    s = {
                        "id": item["id"],
                        "label": (item.get("props") or {}).get("label", "Sub-repetidor"),
                    }
                    for c in item.get("children") or []:
                        _walk(c, rep_ctx, s)
                else:
                    r = {
                        "id": item["id"],
                        "label": (item.get("props") or {}).get("label", "Repetidor"),
                    }
                    for c in item.get("children") or []:
                        _walk(c, r, None)
            elif t not in ("label", "divider", "helpText", "mathoperations", "firm", ""):
                props = item.get("props") or {}
                q_id = str(
                    item.get("id_question")
                    or item.get("linkExternalId")
                    or item.get("id", "")
                )
                fid = item.get("id", "")
                label = props.get("label", "Sin etiqueta")
                opts = props.get("options")

                if sub_ctx and rep_ctx:
                    fields.append(dict(
                        q_id=q_id, fid=fid, type=t, label=label,
                        level="sub-repeater",
                        rep=rep_ctx["id"], sub=sub_ctx["id"],
                        nivel=f'{rep_ctx["id"]}|{sub_ctx["id"]}',
                        opts=opts,
                    ))
                elif rep_ctx:
                    fields.append(dict(
                        q_id=q_id, fid=fid, type=t, label=label,
                        level="repeater",
                        rep=rep_ctx["id"], sub=None,
                        nivel=rep_ctx["id"], opts=opts,
                    ))
                else:
                    fields.append(dict(
                        q_id=q_id, fid=fid, type=t, label=label,
                        level="normal",
                        rep=None, sub=None, nivel="", opts=opts,
                    ))

        for it in fd:
            if it.get("type"):
                _walk(it)

        if not fields:
            raise HTTPException(
                status_code=400,
                detail="No se encontraron campos exportables",
            )

        has_rep = any(f["level"] == "repeater" for f in fields)
        has_sub = any(f["level"] == "sub-repeater" for f in fields)

        # ── 3. Respuestas filtradas por fecha ─────────────────────────────────
        rq = (
            db.query(Response)
            .filter(
                Response.form_id == form_id,
                Response.user_id == current_user.id,
                Response.status != ResponseStatus.draft,
            )
        )
        if date_from:
            try:
                rq = rq.filter(Response.submitted_at >= _dt.fromisoformat(date_from))
            except ValueError:
                raise HTTPException(status_code=400, detail=f"date_from invalido: {date_from}")
        if date_to:
            try:
                rq = rq.filter(Response.submitted_at <= _dt.fromisoformat(date_to))
            except ValueError:
                raise HTTPException(status_code=400, detail=f"date_to invalido: {date_to}")

        resps = rq.order_by(Response.submitted_at.asc()).all()
        if not resps:
            raise HTTPException(
                status_code=404,
                detail="No se encontraron respuestas en el rango indicado",
            )

        rids = [r.id for r in resps]
        all_ans = (
            db.query(Answer)
            .filter(Answer.response_id.in_(rids))
            .order_by(Answer.response_id, Answer.id)
            .all()
        )
        ans_map: dict[int, list] = {}
        for a in all_ans:
            ans_map.setdefault(a.response_id, []).append(a)

        # ── 4. Mapa fid -> col index ─────────────────────────────────────────
        fid_col = {f["fid"]: i for i, f in enumerate(fields)}

        def _col_for(a):
            if a.form_design_element_id and a.form_design_element_id in fid_col:
                return fid_col[a.form_design_element_id]
            for i2, f2 in enumerate(fields):
                if str(a.question_id) == f2["q_id"]:
                    return i2
            return None

        # ── 5. Construir bloques de datos ─────────────────────────────────────
        # Cada Response es un "Envio N".
        #
        # La estructura de filas para repeater replica exactamente la plantilla
        # del frontend (buildFormTemplateWorkbook):
        #   Envio N        -> normal + rep[0] + sub[0]
        #     -> Sub 2     -> sub[1]   (solo celdas sub-repeater)
        #   -> Fila 2      -> rep[1] + sub[0] de fila 1  (celdas grises para normal)
        #     -> Sub 2     -> sub[1] de fila 1
        #   -> Fila 3      -> rep[2] + sub[0] de fila 2

        # IDs de columnas de repeater agrupadas por repeater_id
        rep_field_indices: dict[str, list[int]] = {}  # rep_id -> [col_indices]
        sub_field_indices: dict[str, list[int]] = {}  # sub_id -> [col_indices]
        for i, f in enumerate(fields):
            if f["level"] == "repeater" and f["rep"]:
                rep_field_indices.setdefault(f["rep"], []).append(i)
            elif f["level"] == "sub-repeater" and f["sub"]:
                sub_field_indices.setdefault(f["sub"], []).append(i)

        blocks: list[tuple] = []

        for resp in resps:
            ra = ans_map.get(resp.id, [])

            if not has_rep and not has_sub:
                rd = [""] * len(fields)
                for a in ra:
                    c = _col_for(a)
                    if c is not None:
                        rd[c] = a.answer_text or ""
                blocks.append(("simple", [rd]))
                continue

            # ── Clasificar answers de esta respuesta ──────────────────────
            normal_vals = [""] * len(fields)
            # repeater answers agrupadas por columna
            rep_by_col: dict[int, list] = {}   # col_index -> [answer objects ordered]
            sub_by_col: dict[int, list] = {}

            for a in ra:
                c = _col_for(a)
                if c is None:
                    continue
                f = fields[c]
                if f["level"] == "normal":
                    normal_vals[c] = a.answer_text or ""
                elif f["level"] == "repeater":
                    rep_by_col.setdefault(c, []).append(a)
                elif f["level"] == "sub-repeater":
                    sub_by_col.setdefault(c, []).append(a)

            # Ordenar cada columna de repeater por repeater_row_index (si
            # existe) o por id (orden de insercion).
            for col_answers in rep_by_col.values():
                has_idx = any(a.repeater_row_index is not None for a in col_answers)
                if has_idx:
                    col_answers.sort(key=lambda a: (a.repeater_row_index or 0, a.id))
                else:
                    col_answers.sort(key=lambda a: a.id)

            for col_answers in sub_by_col.values():
                has_idx = any(a.repeater_row_index is not None for a in col_answers)
                if has_idx:
                    col_answers.sort(key=lambda a: (
                        a.parent_row_index or 0,
                        a.repeater_row_index or 0,
                        a.id,
                    ))
                else:
                    col_answers.sort(key=lambda a: a.id)

            # ── Calcular cuantas filas de repeater hay ────────────────────
            # = max answers de cualquier columna de repeater
            max_rep_rows = 0
            for col_answers in rep_by_col.values():
                max_rep_rows = max(max_rep_rows, len(col_answers))
            if max_rep_rows == 0:
                max_rep_rows = 1  # al menos la fila principal

            # ── Calcular sub-repeater filas por fila padre ────────────────
            # sub_rows_per_parent[parent_row_idx] = max sub rows
            sub_rows_per_parent: dict[int, int] = {}

            for ci, col_answers in sub_by_col.items():
                has_pidx = any(a.parent_row_index is not None for a in col_answers)
                if has_pidx:
                    # Agrupar por parent_row_index
                    by_parent: dict[int, int] = {}
                    for a in col_answers:
                        pi = a.parent_row_index or 0
                        by_parent[pi] = by_parent.get(pi, 0) + 1
                    for pi, cnt in by_parent.items():
                        sub_rows_per_parent[pi] = max(
                            sub_rows_per_parent.get(pi, 0), cnt
                        )
                else:
                    # Sin parent_row_index: dividir equitativamente
                    total_sub = len(col_answers)
                    per_parent = max(1, total_sub // max_rep_rows) if max_rep_rows > 0 else total_sub
                    for pi in range(max_rep_rows):
                        sub_rows_per_parent[pi] = max(
                            sub_rows_per_parent.get(pi, 0), per_parent
                        )

            # ── Construir filas del bloque ─────────────────────────────────
            brows: list[tuple] = []

            for ri in range(max_rep_rows):
                # -- Fila principal/repeater --
                row = [""] * len(fields)

                # Campos normales solo en fila 0
                if ri == 0:
                    for i, v in enumerate(normal_vals):
                        if fields[i]["level"] == "normal":
                            row[i] = v

                # Campos de repeater: tomar el answer [ri] de cada columna
                for ci, col_answers in rep_by_col.items():
                    if ri < len(col_answers):
                        row[ci] = col_answers[ri].answer_text or ""

                # Sub-repeater fila 0 de esta fila padre
                for ci, col_answers in sub_by_col.items():
                    has_pidx = any(a.parent_row_index is not None for a in col_answers)
                    if has_pidx:
                        match = [a for a in col_answers
                                 if (a.parent_row_index or 0) == ri
                                 and (a.repeater_row_index or 0) == 0]
                        if match:
                            row[ci] = match[0].answer_text or ""
                    else:
                        # Sin parent_row_index: calcular offset
                        per_parent = sub_rows_per_parent.get(ri, 0)
                        idx = ri * per_parent
                        if idx < len(col_answers):
                            row[ci] = col_answers[idx].answer_text or ""

                brows.append(("main" if ri == 0 else "rep", ri, row))

                # -- Sub-repeater filas adicionales (sr >= 1) --
                n_sub = sub_rows_per_parent.get(ri, 0)
                for si in range(1, n_sub):
                    sr_row = [""] * len(fields)
                    for ci, col_answers in sub_by_col.items():
                        has_pidx = any(a.parent_row_index is not None for a in col_answers)
                        if has_pidx:
                            match = [a for a in col_answers
                                     if (a.parent_row_index or 0) == ri
                                     and (a.repeater_row_index or 0) == si]
                            if match:
                                sr_row[ci] = match[0].answer_text or ""
                        else:
                            per_parent = sub_rows_per_parent.get(ri, 0)
                            idx = ri * per_parent + si
                            if idx < len(col_answers):
                                sr_row[ci] = col_answers[idx].answer_text or ""
                    brows.append(("sub", si, sr_row))

            blocks.append(("repeater", brows))

        # ── 6. Generar Excel con openpyxl ─────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla"

        tb = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        hfont = Font(bold=True, size=11, color="FFFFFF")
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        lfill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        mfont = Font(size=8, color="AAAAAA")
        mfill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        dfont = Font(size=10)
        gfill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        sfill = PatternFill(start_color="F0FDFA", end_color="F0FDFA", fill_type="solid")
        sfont = Font(italic=True, size=9, color="374151")
        ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
        la = Alignment(horizontal="left", vertical="center", wrap_text=True)

        CO = 2  # col offset: A=label, B=sep, C+=campos

        def _wh(r, c, v):
            x = ws.cell(row=r, column=c, value=v)
            x.font = hfont; x.fill = hfill; x.alignment = ca; x.border = tb

        def _wl(r, c, v):
            x = ws.cell(row=r, column=c, value=v)
            x.font = Font(bold=True, size=10); x.fill = lfill; x.alignment = la; x.border = tb

        def _wm(r, c, v):
            x = ws.cell(row=r, column=c, value=v)
            x.font = mfont; x.fill = mfill; x.alignment = la; x.border = tb

        def _wd(r, c, v=""):
            x = ws.cell(row=r, column=c, value=v)
            x.font = dfont; x.alignment = la; x.border = tb

        def _wg(r, c):
            x = ws.cell(row=r, column=c, value="")
            x.font = dfont; x.fill = gfill; x.alignment = ca; x.border = tb

        def _wsub(r, c, v):
            x = ws.cell(row=r, column=c, value=v)
            x.font = sfont; x.fill = sfill; x.alignment = la; x.border = tb

        # ── Filas de metadatos (1-4) ──────────────────────────────────────────
        _wl(1, 1, "ID Pregunta"); _wd(1, 2, "")
        _wl(2, 1, "Tipo");        _wd(2, 2, "")
        _wm(3, 1, "Nivel");       _wm(3, 2, "")
        _wl(4, 1, "Pregunta");    _wd(4, 2, "")

        for i, f in enumerate(fields):
            c = i + 1 + CO
            _wh(1, c, f["q_id"])
            _wh(2, c, f["type"])
            _wm(3, c, f["nivel"])
            _wh(4, c, f["label"])

        # ── Filas de datos (5+) ───────────────────────────────────────────────
        cr = 5
        for bi, (btype, bdata) in enumerate(blocks):
            en = bi + 1

            if btype == "simple":
                rd = bdata[0]
                _wl(cr, 1, f"Env\u00edo {en}")
                _wd(cr, 2, "")
                for i, v in enumerate(rd):
                    _wd(cr, i + 1 + CO, v)
                cr += 1
            else:
                # repeater block
                for rtype, ridx, rdata in bdata:
                    if rtype == "main":
                        _wl(cr, 1, f"Env\u00edo {en}")
                    elif rtype == "rep":
                        _wsub(cr, 1, f"  \u21b3 Fila {ridx + 1}")
                    elif rtype == "sub":
                        _wsub(cr, 1, f"    \u21b3 Sub {ridx + 1}")

                    _wd(cr, 2, "")

                    for i, v in enumerate(rdata):
                        fl = fields[i]["level"]
                        if rtype == "sub" and fl != "sub-repeater":
                            _wg(cr, i + 1 + CO)
                        elif rtype == "rep" and fl == "normal":
                            _wg(cr, i + 1 + CO)
                        else:
                            _wd(cr, i + 1 + CO, v)
                    cr += 1

        # ── Anchos de columna ─────────────────────────────────────────────────
        ws.column_dimensions[get_column_letter(1)].width = 18
        ws.column_dimensions[get_column_letter(2)].width = 3
        for i in range(len(fields)):
            ws.column_dimensions[get_column_letter(i + 1 + CO)].width = 28

        # ── Alturas de fila metadata ──────────────────────────────────────────
        ws.row_dimensions[1].height = 16
        ws.row_dimensions[2].height = 14
        ws.row_dimensions[3].height = 11
        ws.row_dimensions[4].height = 55

        # ── Data validation para select/radio ─────────────────────────────────
        for i, f in enumerate(fields):
            o = f.get("opts")
            if o and isinstance(o, list) and len(o) > 0 and f["type"] in ("select", "radio"):
                os_str = ",".join(str(x) for x in o)
                if len(os_str) <= 250:
                    cl = get_column_letter(i + 1 + CO)
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{os_str}"',
                        allow_blank=True,
                        showInputMessage=True,
                        showErrorMessage=True,
                        promptTitle=f["label"],
                        prompt="Seleccione una opcion",
                        errorTitle="Valor no valido",
                        error="Seleccione un valor de la lista.",
                    )
                    dv.add(f"{cl}5:{cl}{cr}")
                    ws.add_data_validation(dv)

        # ── Guardar y retornar ────────────────────────────────────────────────
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_title = form.title.replace(" ", "_").replace("/", "_")[:80]
        filename = f"{safe_title}_respuestas_plantilla.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
