from datetime import datetime
import logging
import math
import os
from pathlib import Path
import shutil
import uuid
from fastapi import APIRouter, Depends, HTTPException, UploadFile,Query,Request,  File, Body, status, Form as FastAPIForm
from sqlalchemy import and_, desc, func, or_, select
from sqlalchemy.orm import Session, joinedload, defer
from typing import Any, List, Optional
from app.api.controllers.excel_form_exporter import generate_form_excel
from app.api.controllers.mail import send_response_answers_email
from app.redis_client import redis_client
from app.database import get_db
from app.models import Answer, AnswerHistory, ApprovalStatus, CategoryApproval, Form, FormAnswer, FormApproval, FormApprovalNotification, FormCategory, FormCloseConfig, FormModerators, FormMovimientos, FormQuestion, FormSchedule, FormTemplate, PalabrasClave, Question, QuestionTableRelation, QuestionType, Response, ResponseApproval, ResponseStatus, TemplateScope, User, UserType
from app.crud import  _extract_style_config, _serialize_answers, add_category_approver, analyze_form_relations, apply_template_service, bulk_save_category_approvers, check_form_data, create_form, add_questions_to_form, create_form_category, create_form_movimiento, create_form_schedule, create_response_approval, create_template_service, delete_form, delete_form_category, delete_template_service, fetch_completed_forms_by_user, fetch_completed_forms_with_all_responses, fetch_form_questions, fetch_form_users, generate_excel_with_repeaters, get_all_categories_with_approvers, get_all_form_movimientos_basic, get_all_forms, get_all_forms_paginated, get_all_user_responses_by_form_id_improved, get_categories_by_parent, get_category_approvals, get_category_path, get_category_tree, get_form, get_form_id_users, get_form_responses_data, get_form_with_full_responses, get_forms, get_forms_by_approver, get_forms_by_user, get_forms_by_user_summary, get_forms_pending_approval_for_user, get_moderated_forms_by_answers, get_next_mandatory_approver, get_notifications_for_form, get_questions_and_answers_by_form_id, get_questions_and_answers_by_form_id_and_user, get_response_approval_status, get_response_details_logic, get_template_detail_service, get_unanswered_forms_by_user, get_user_responses_data, invalidate_form_cache, link_moderator_to_form, link_question_to_form, list_templates_service, move_category, process_regisfacial_answer, remove_category_approver, remove_moderator_from_form, remove_question_from_form, save_form_approvals, search_forms_by_user, send_rejection_email_to_all, sync_form_approvals_from_category, toggle_form_status, update_category_approver, update_form_category_1, update_form_design_service, update_notification_status, update_response_approval_status, update_template_service
from app.schemas import AlertMessageRequest, CategoryApprovalBulkSave, CategoryApprovalCreate, CategoryApprovalResponse, CategoryApprovalUpdate, FormAnswerCreate, FormBaseUser, FormCategoryCreate, FormCategoryMove, FormCategoryResponse, FormCategoryTreeResponse, FormCategoryUpdate, FormCategoryWithFormsResponse, FormCloseConfigCreate, FormCloseConfigOut, FormCreate, FormDesignUpdate, FormMovimientoBase, FormMovimientoResponse, FormResponse, FormResponseBitacora, FormScheduleCreate, FormScheduleOut, FormStatusUpdate, FormTemplateCreate, FormTemplateDetail, FormTemplateResponse, FormTemplateUpdate, NotificationCreate, NotificationsByFormResponse_schema, QuestionAdd, FormBase, QuestionIdsRequest, RelatedAnswerRequest, ResponseApprovalCreate, SendResponseEmailRequest, UpdateFormBasicInfo, UpdateFormCategory, UpdateNotifyOnSchema, UpdateResponseApprovalRequest
from app.core.security import get_current_user
from io import BytesIO
import pandas as pd
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse

router = APIRouter()

MAX_APPROVALS_PER_FORM = 15

@router.post("/", response_model=FormResponse, status_code=status.HTTP_201_CREATED)
def create_form_endpoint(
    form: FormBaseUser,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Crear un nuevo formulario.

    Este endpoint permite a los usuarios con el rol de `creator` o `admin` crear
    un nuevo formulario. Los usuarios que no tengan estos permisos recibirán
    un error HTTP 403 (Prohibido).

    Args:
        form (FormBaseUser): Los datos del formulario que se va a crear.
        db (Session): Sesión de base de datos proporcionada por la dependencia.
        current_user (User): Usuario autenticado extraído del token JWT.

    Returns:
        FormResponse: Objeto con la información del formulario creado.

    Raises:
        HTTPException: Si el usuario no tiene permisos adecuados para crear formularios.
    """
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )

    return create_form(db=db, form=form, user_id=current_user.id)

@router.get("/category-approvals", summary="Todas las categorías con sus aprobadores")
def list_all_categories_with_approvers(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Devuelve todas las categorías con sus aprobadores asignados."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    return get_all_categories_with_approvers(db)


@router.put("/category-approvals/approver/{approval_id}", response_model=CategoryApprovalResponse)
def update_approver(
    approval_id: int,
    payload: CategoryApprovalUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Actualiza un aprobador (secuencia, obligatoriedad, plazo)."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    return update_category_approver(db, approval_id, payload)


@router.delete("/category-approvals/approver/{approval_id}", status_code=200)
def delete_approver(
    approval_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Elimina un aprobador de una categoría."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    remove_category_approver(db, approval_id)
    return {"message": "Aprobador eliminado correctamente"}


# ⚠️ RUTAS DINÁMICAS DESPUÉS

@router.get("/category-approvals/{category_id}", response_model=List[CategoryApprovalResponse])
def get_approvers_by_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtiene los aprobadores activos de una categoría específica."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    return get_category_approvals(db, category_id)

@router.get("/categories/{category_id}/has-approvers")
def check_category_approvers(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Verifica si una categoría tiene aprobadores definidos."""
    approvers = db.query(CategoryApproval).options(
        joinedload(CategoryApproval.user)
    ).filter(
        CategoryApproval.category_id == category_id,
        CategoryApproval.is_active == True
    ).order_by(CategoryApproval.sequence_number).all()

    return {
        "has_approvers": len(approvers) > 0,
        "count": len(approvers),
        "approvers": [
            {
                "user_id": a.user_id,
                "user_name": a.user.name if a.user else f"Usuario #{a.user_id}",
                "sequence_number": a.sequence_number,
                "is_mandatory": a.is_mandatory,
            }
            for a in approvers
        ]
    }

@router.post("/category-approvals/{category_id}", response_model=CategoryApprovalResponse, status_code=201)
def add_approver_to_category(
    category_id: int,
    payload: CategoryApprovalCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Agrega un aprobador a una categoría."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    return add_category_approver(db, category_id, payload)


@router.put("/category-approvals/{category_id}/bulk", response_model=List[CategoryApprovalResponse])
def bulk_save_approvers(
    category_id: int,
    payload: CategoryApprovalBulkSave,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Guarda toda la lista de aprobadores de una categoría (reemplaza existentes)."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    return bulk_save_category_approvers(db, category_id, payload.approvers)


@router.post("/form-templates", response_model=FormTemplateResponse, status_code=201)
def create_template(
    payload: FormTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Crea una nueva plantilla de formulario."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    return create_template_service(db, current_user.id, payload)


@router.get("/form-templates", response_model=List[FormTemplateResponse])
def list_templates(
    id_category: Optional[int] = Query(None, description="Filtrar por ID de categoría"),
    search: Optional[str] = Query(None, description="Buscar por nombre o descripción"),
    scope: Optional[str] = Query(None, description="Filtrar por alcance: private, company, public"),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Lista plantillas accesibles para el usuario actual."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    return list_templates_service(db, current_user.id, id_category, search, scope, skip, limit)


@router.get("/form-templates/categories")
def list_template_categories(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Devuelve todas las categorías disponibles (las mismas de formularios)."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")

    categories = db.query(FormCategory).order_by(FormCategory.order, FormCategory.name).all()
    return [
        {
            "id": cat.id,
            "name": cat.name,
            "description": cat.description,
            "icon": cat.icon,
            "color": cat.color,
            "parent_id": cat.parent_id,
        }
        for cat in categories
    ]


@router.get("/form-templates/{template_id}", response_model=FormTemplateDetail)
def get_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtiene el detalle completo de una plantilla incluyendo su diseño."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    return get_template_detail_service(db, template_id, current_user.id)


@router.post("/form-templates/{template_id}/apply")
def apply_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Aplica una plantilla: devuelve diseño con UUIDs frescos."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    design = apply_template_service(db, template_id, current_user.id)
    return {"template_design": design}


@router.put("/form-templates/{template_id}", response_model=FormTemplateResponse)
def update_template(
    template_id: int,
    payload: FormTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Actualiza metadatos o diseño de una plantilla propia."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    return update_template_service(db, template_id, current_user.id, payload)


@router.delete("/form-templates/{template_id}", status_code=200)
def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Elimina (soft-delete) una plantilla propia."""
    if current_user is None:
        raise HTTPException(status_code=403, detail="Authentication required")
    delete_template_service(db, template_id, current_user.id)
    return {"message": "Template deleted successfully"}


@router.post("/{form_id}/questions", response_model=FormResponse)
def add_questions_to_form_endpoint(
    form_id: int,
    questions: QuestionAdd,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Agrega preguntas a un formulario existente.

    Este endpoint permite agregar una o más preguntas a un formulario específico.
    Solo los usuarios con rol `creator` o `admin` tienen permiso para realizar esta acción.

    Args:
        form_id (int): ID del formulario al cual se desean agregar las preguntas.
        questions (QuestionAdd): Objeto que contiene una lista de IDs de preguntas a agregar.
        db (Session): Sesión activa de la base de datos, inyectada por dependencia.
        current_user (User): Usuario autenticado actual, obtenido desde el token JWT.

    Returns:
        FormResponse: Objeto que representa el formulario actualizado con las nuevas preguntas.

    Raises:
        HTTPException: Error 403 si el usuario no tiene permisos suficientes para modificar el formulario.
    """
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )
    return add_questions_to_form(db, form_id, questions.question_ids)

@router.get("/all")
def get_all_forms_endpoint(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):   
    """
    Obtener todos los formularios del usuario autenticado.

    Este endpoint recupera la información completa de todos los formularios 
    asociados al usuario autenticado, incluyendo preguntas, opciones y respuestas.

    Args:
        db (Session): Sesión activa de la base de datos, inyectada como dependencia.
        current_user (User): Usuario autenticado, obtenido desde el token JWT.

    Returns:
        list: Lista de objetos con los datos de todos los formularios del usuario.
    """
    # Obtener todos los form_ids del usuario
    user_forms = db.query(Form.id).filter(Form.user_id == current_user.id).all()
    
    if not user_forms:
        return []
    
    # Obtener los detalles completos de cada formulario
    forms_data = []
    for form_tuple in user_forms:
        form_id = form_tuple[0]
        form_detail = get_form(db, form_id, current_user.id)
        if form_detail:
            forms_data.append(form_detail)
    
    return forms_data

@router.get("/list")
def get_forms_list(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener solo id y título de los formatos del usuario autenticado.
    """
    forms = (
        db.query(Form.id, Form.title)
        .order_by(Form.created_at.desc())
        .all()
    )

    return [
        {
            "id": form.id,
            "title": form.title
        }
        for form in forms
    ]


@router.get("/{form_id}")
def get_form_endpoint(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):   
    """
    Obtener un formulario específico por su ID.

    Este endpoint recupera la información de un formulario asociado al usuario autenticado.
    Si el formulario no existe o no pertenece al usuario, se devuelve un error 404.

    Args:
        form_id (int): ID del formulario que se desea consultar.
        db (Session): Sesión activa de la base de datos, inyectada como dependencia.
        current_user (User): Usuario autenticado, obtenido desde el token JWT.

    Returns:
        dict: Objeto con los datos del formulario solicitado.

    Raises:
        HTTPException: Error 404 si el formulario no se encuentra o no pertenece al usuario.
    """
    form = get_form_id_users(db, form_id, current_user.id)
    if not form:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Form not found")
    return form


@router.get("/{form_id}/form_design")
def get_form_design(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener solo el diseño visual de un formulario (sin respuestas).
    
    Este endpoint es ultra-ligero y usa caché de Redis con TTL de 1 hora.
    Ideal para renderizar la UI del formulario sin cargar respuestas.
    
    Returns:
        dict: Diseño del formulario (id, title, description, version, form_design)
    """
    # PASO 1: Verificar caché Redis
    cache_key = f"form_design:{form_id}"
    cached = redis_client.get(cache_key)  # Ahora usa tu método .get()
    
    if cached:
        print(f"✅ Cache HIT: {cache_key}")
        return cached  # Ya viene deserializado por tu método
    
    # PASO 2: Cache MISS - Consultar BD
    print(f"❌ Cache MISS: {cache_key}")
    form = db.query(Form).filter(Form.id == form_id).first()
    
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    # PASO 3: Extraer solo el diseño
    design_response = {
        "id": form.id,
        "title": form.title,
        "description": form.description,
        "created_at": form.created_at.isoformat() if hasattr(form, 'created_at') and form.created_at else None,
        "format_type": form.format_type.name if hasattr(form, 'format_type') else None,
        "form_design": form.form_design  # Componentes visuales
    }
    
    # PASO 4: Guardar en Redis (TTL: 1 hora = 3600 segundos)
    redis_client.set(cache_key, design_response, ttl=3600)  # Usa tu método .set()
    
    return design_response

@router.get("/{form_id}/questions")
def get_form_questions(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener solo la metadata de preguntas (sin respuestas).
    
    Incluye: id, texto, tipo, opciones, validaciones, orden.
    Caché de 1 hora en Redis.
    
    Returns:
        dict: Metadata de todas las preguntas del formulario
    """
    # PASO 1: Verificar caché Redis
    cache_key = f"form_questions:{form_id}"
    cached = redis_client.get(cache_key)
    
    if cached:
        print(f"✅ Cache HIT: {cache_key}")
        return cached
    
    # PASO 2: Consultar BD (solo questions)
    print(f"❌ Cache MISS: {cache_key}")
    
    # Obtener el formulario primero
    form = db.query(Form).filter(Form.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    # Obtener preguntas asociadas al formulario
    form_questions = (
        db.query(FormQuestion)
        .filter(FormQuestion.form_id == form_id)
        .all()
    )
    
    if not form_questions:
        raise HTTPException(status_code=404, detail="No questions found for this form")
    
    # Obtener los IDs de las preguntas
    question_ids = [fq.question_id for fq in form_questions]
    
    # Obtener las preguntas con sus opciones
    questions = (
        db.query(Question)
        .filter(Question.id.in_(question_ids))
        .options(joinedload(Question.options))
        .all()
    )
    
    if not questions:
        raise HTTPException(status_code=404, detail="Questions not found") 
    form = db.query(Form).filter(Form.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    # Obtener preguntas asociadas al formulario
    form_questions = (
        db.query(FormQuestion)
        .filter(FormQuestion.form_id == form_id)
        .all()
    )
    
    if not form_questions:
        # Si no hay FormQuestion, intentar obtener directamente de Question
        questions = (
            db.query(Question)
            .join(FormQuestion, FormQuestion.question_id == Question.id)
            .filter(FormQuestion.form_id == form_id)
            .options(joinedload(Question.options))
            .all()
        )
        
        if not questions:
            raise HTTPException(status_code=404, detail="No questions found for this form")
        
        questions_map = {q.id: q for q in questions}
    else:
        # Obtener los IDs de las preguntas
        question_ids = [fq.question_id for fq in form_questions]
        
        # Obtener las preguntas con sus opciones
        questions = (
            db.query(Question)
            .filter(Question.id.in_(question_ids))
            .options(joinedload(Question.options))
            .all()
        )
        
        # Crear un mapa de preguntas por ID
        questions_map = {q.id: q for q in questions}
    
    # PASO 3: Obtener globalConfig para resolver optionSets
    global_config = {}
    if hasattr(form, 'global_config') and form.global_config:
        global_config = form.global_config
    option_sets = global_config.get("optionSets", {})
    
    # PASO 4: Preparar respuesta (resolver referencias a optionSets)
    questions_response = {
        "form_id": form_id,
        "optionSets": option_sets,
        "questions": []
    }
    
    # Procesar cada pregunta
    for question in questions:
        question_data = {
            "id": question.id,
            "question_text": question.question_text,
            "question_type": question.question_type,
            "required": question.required,
            "root": question.root if hasattr(question, 'root') else False,
            "options": [
                {"id": opt.id, "option_text": opt.option_text}
                for opt in question.options
            ]
        }
        questions_response["questions"].append(question_data)
    
    # PASO 5: Guardar en Redis (TTL: 1 hora)
    redis_client.set(cache_key, questions_response, ttl=3600)
    
    return questions_response


@router.get("/{form_id}/responses/user")
def get_user_responses(
    form_id: int,
    response_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener solo las respuestas del usuario autenticado filtradas por response_id.
    
    Este endpoint usa caché porque las respuestas del usuario no cambian con frecuencia.
    Optimizado para ser ultra-ligero y rápido.
    
    Args:
        form_id (int): ID del formulario
        response_id (int): ID de la respuesta específica
        
    Returns:
        dict: Respuestas del usuario al formulario
    """
    # PASO 1: Verificar caché Redis
    cache_key = f"user_responses:{form_id}:{response_id}:{current_user.id}"
    cached = redis_client.get(cache_key)
    
    if cached:
        print(f"✅ Cache HIT: {cache_key}")
        return cached
    
    # PASO 2: Consultar BD (con joinedload optimizado + filtro por response_id)
    print(f"❌ Cache MISS: {cache_key}")
    
    responses = (
        db.query(Response)
        .filter(
            Response.form_id == form_id,
            Response.id == response_id,
            Response.user_id == current_user.id
        )
        .options(joinedload(Response.answers))
        .all()
    )
    
    if not responses:
        raise HTTPException(
            status_code=404, 
            detail="No se encontró la respuesta especificada"
        )
    
    # PASO 3: Obtener preguntas UNA VEZ (para procesar regisfacial)
    questions = db.query(Question).filter(
        Question.id.in_([
            answer.question_id 
            for response in responses 
            for answer in response.answers
        ])
    ).all()
    questions_map = {q.id: q for q in questions}
    
    # PASO 4: Procesar respuestas (incluyendo regisfacial)
    result = []
    for response in responses:
        response_data = {
            "id": response.id,
            "form_id": response.form_id,
            "submitted_at": response.submitted_at.isoformat() if response.submitted_at else None,
            "status": response.status if hasattr(response, 'status') else "completed",
            "answers": []
        }
        
        for answer in response.answers:
            question = questions_map.get(answer.question_id)
            answer_text = answer.answer_text
            
            # Procesar regisfacial
            if question and question.question_type == "regisfacial":
                answer_text = process_regisfacial_answer(answer_text, "regisfacial")
            
            response_data["answers"].append({
                "question_id": answer.question_id,
                "answer_text": answer_text
            })
        
        result.append(response_data)
    
    # PASO 5: Guardar en Redis (TTL: 1 hora - raramente cambia)
    redis_client.set(cache_key, result, ttl=3600)
    
    return result

@router.get("/{form_id}/has-responses")
def check_form_responses(form_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Verifica si un formulario tiene respuestas asociadas y retorna sus datos completos.

    Este endpoint permite obtener información detallada sobre un formulario específico,
    incluyendo si tiene respuestas, las respuestas en sí, los usuarios que respondieron
    y las preguntas con sus respectivas respuestas. Solo los usuarios autenticados
    pueden acceder a esta información.

    Args:
        form_id (int): ID del formulario a consultar.
        db (Session): Sesión de la base de datos SQLAlchemy, proporcionada automáticamente.
        current_user (User): Usuario autenticado, inyectado desde el token de sesión.

    Returns:
        dict: Objeto con los datos del formulario, su creador, proyecto, preguntas y respuestas.

    Raises:
        HTTPException: 
            - 403 si el usuario no está autenticado.
            - 404 si el formulario no existe.
    """
    if current_user == None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to get form"
        )
    else:    
        return check_form_data(db, form_id)
    

@router.put("/{form_id}/questions")
async def update_form_questions(
    form_id: int,
    request: QuestionIdsRequest,  # { "question_ids": [1, 2, 3] }
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Actualiza las relaciones de preguntas para un formulario (reemplaza las existentes)
    Útil para modo edición del formulario
    """
    try:
        # Verificar que el formulario existe
        form = db.query(Form).filter(Form.id == form_id).first()
        
        if not form:
            raise HTTPException(status_code=404, detail="Formulario no encontrado")
        
        # Eliminar todas las relaciones existentes
        db.query(FormQuestion).filter(FormQuestion.form_id == form_id).delete()
        
        # Filtrar solo las preguntas que existen
        valid_question_ids = []
        invalid_question_ids = []
        
        if request.question_ids and len(request.question_ids) > 0:
            for question_id in request.question_ids:
                # Verificar que la pregunta existe
                question = db.query(Question).filter(Question.id == question_id).first()
                
                if question:
                    # Crear la relación solo si la pregunta existe
                    form_question = FormQuestion(
                        form_id=form_id,
                        question_id=question_id
                    )
                    db.add(form_question)
                    valid_question_ids.append(question_id)
                else:
                    # Registrar las preguntas inválidas para informar al usuario
                    invalid_question_ids.append(question_id)
        
        db.commit()
        
        # 🔥 INVALIDAR CACHÉ DE REDIS
        invalidate_form_cache(form_id)
        
        response = {
            "success": True,
            "message": "Preguntas actualizadas correctamente",
            "form_id": form_id,
            "valid_question_ids": valid_question_ids
        }
        
        # Informar si hubo preguntas que no se pudieron agregar
        if invalid_question_ids:
            response["warning"] = f"Las siguientes preguntas no existen y fueron omitidas: {invalid_question_ids}"
            response["invalid_question_ids"] = invalid_question_ids
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    
    
@router.get("/emails/all-emails")
def get_all_emails(db: Session = Depends(get_db)):
    """
    Obtiene todos los correos electrónicos de los usuarios registrados.

    Este endpoint recupera todos los correos electrónicos de los usuarios en la base de datos
    y los devuelve como una lista. No requiere autenticación, pero dependiendo del caso de uso
    se recomienda proteger este endpoint para evitar exposición de datos sensibles.

    Args:
        db (Session): Sesión de la base de datos, inyectada automáticamente por FastAPI.

    Returns:
        dict: Un diccionario con la clave `"emails"` que contiene una lista de correos electrónicos.

    Ejemplo de respuesta:
    {
        "emails": [
            "usuario1@example.com",
            "usuario2@example.com",
            ...
        ]
    }
    """
    emails = db.query(User.email).all()
    return {"emails": [email[0] for email in emails]}




@router.post("/form_schedules/")
def register_form_schedule(schedule_data: FormScheduleCreate, db: Session = Depends(get_db)):
    """
    Registra o actualiza la programación de un formulario.

    Este endpoint permite crear o actualizar la programación automática de envío o visualización 
    de un formulario para un usuario específico. Si ya existe una programación con la misma 
    combinación de `form_id` y `user_id`, se actualizará con los nuevos valores. De lo contrario, 
    se creará un nuevo registro.

    Args:
        schedule_data (FormScheduleCreate): Datos para registrar o actualizar la programación del formulario.
        db (Session): Sesión de base de datos proporcionada por FastAPI.

    Returns:
        FormSchedule: Objeto de programación recién creado o actualizado.

    """
    return create_form_schedule(
        db=db,
        form_id=schedule_data.form_id,
        user_id=schedule_data.user_id,
        frequency_type=schedule_data.frequency_type,
        repeat_days=schedule_data.repeat_days,
        interval_days=schedule_data.interval_days,
        specific_date=schedule_data.specific_date,
        status=schedule_data.status
    )

import json

@router.get("/responses/")
def get_responses_with_answers(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Obtiene todas las respuestas completadas por el usuario autenticado para un formulario específico,
    incluyendo sus respuestas, aprobaciones y estado de revisión.
    Maneja el historial de respuestas para mostrar solo las más recientes.

    Args:
        form_id (int): ID del formulario del cual se desean obtener las respuestas.
        db (Session): Sesión activa de la base de datos.
        current_user (User): Usuario autenticado.

    Returns:
        List[dict]: Lista de respuestas con sus respectivos detalles de aprobación y respuestas a preguntas.

    Raises:
        HTTPException: 403 si no hay un usuario autenticado.
        HTTPException: 404 si no se encuentran respuestas.
    """
    if not current_user:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to access completed forms",
        )

    stmt = (
        select(Response)
        .where(Response.form_id == form_id, Response.user_id == current_user.id)
        .options(
            joinedload(Response.answers).joinedload(Answer.question),
            joinedload(Response.approvals).joinedload(ResponseApproval.user)
        )
    )

    responses = db.execute(stmt).unique().scalars().all()

    if not responses:
        raise HTTPException(status_code=404, detail="No se encontraron respuestas")

    # Función auxiliar para procesar respuestas de reconocimiento facial
    def process_regisfacial_answer(answer_text, question_type):
        """
        Procesa las respuestas de tipo regisfacial para mostrar un texto descriptivo
        del registro facial guardado en lugar del JSON completo de faceData
        """
        if question_type != "regisfacial" or not answer_text:
            return answer_text
        
        try:
            # Intentar parsear el JSON
            face_data = json.loads(answer_text)
            
            # Buscar en diferentes estructuras posibles
            person_name = "Usuario"
            success = False
            
            # Estructura 1: {"faceData": {"success": true, "personName": "..."}}
            if isinstance(face_data, dict) and "faceData" in face_data:
                face_info = face_data["faceData"]
                if isinstance(face_info, dict):
                    success = face_info.get("success", False)
                    person_name = face_info.get("personName", "Usuario")
            
            # Estructura 2: directamente {"success": true, "personName": "..."}
            elif isinstance(face_data, dict):
                success = face_data.get("success", False)
                person_name = face_data.get("personName", face_data.get("person_name", "Usuario"))
            
            # Buscar también otras variantes de nombres
            if person_name == "Usuario":
                person_name = face_data.get("name", face_data.get("user_name", "Usuario"))
            
            if success:
                return f"Datos biométricos de {person_name} registrados"
            else:
                return f"Error en el registro de datos biométricos de {person_name}"
            
        except (json.JSONDecodeError, KeyError, TypeError):
            # Si hay error al parsear JSON, devolver un mensaje genérico
            return "Datos biométricos procesados"

    # Obtener todos los response_ids para buscar el historial
    response_ids = [response.id for response in responses]
    
    # Obtener historiales de respuestas
    histories = db.query(AnswerHistory).filter(AnswerHistory.response_id.in_(response_ids)).all()
    
    # Obtener todos los IDs de respuestas (previous y current) del historial
    all_answer_ids = set()
    for history in histories:
        if history.previous_answer_id:
            all_answer_ids.add(history.previous_answer_id)
        all_answer_ids.add(history.current_answer_id)
    
    # Obtener todas las respuestas del historial con sus preguntas
    historical_answers = {}
    if all_answer_ids:
        historical_answer_list = (
            db.query(Answer)
            .options(joinedload(Answer.question))
            .filter(Answer.id.in_(all_answer_ids))
            .all()
        )
        
        # Crear mapeo de answer_id -> Answer
        for answer in historical_answer_list:
            historical_answers[answer.id] = answer
    
    # Crear mapeo de current_answer_id -> history
    history_map = {}
    # Crear conjunto de previous_answer_ids para saber cuáles no mostrar individualmente
    previous_answer_ids = set()
    
    for history in histories:
        history_map[history.current_answer_id] = history
        if history.previous_answer_id:
            previous_answer_ids.add(history.previous_answer_id)

    result = []
    for r in responses:
        approval_result = get_response_approval_status(r.approvals)

        # Obtener respuestas actuales (excluyendo las que son previous_answer_ids)
        current_answers = []
        for answer in r.answers:
            # Solo incluir respuestas que no sean previous_answer_ids (es decir, las más recientes)
            if answer.id not in previous_answer_ids:
                current_answers.append(answer)

        result.append({
            "response_id": r.id,
            "status": r.status,
            "submitted_at": r.submitted_at,
            "approval_status": approval_result["status"],
            "message": approval_result["message"],
            "answers": [
                {
                    "id_answer": a.id,
                    "repeated_id": r.repeated_id,
                    "question_id": a.question.id,
                    "question_text": a.question.question_text,
                    "question_type": a.question.question_type,
                    "answer_text": process_regisfacial_answer(a.answer_text, a.question.question_type),
                    "file_path": a.file_path,
                    "form_design_element_id": a.form_design_element_id
                }
                for a in current_answers
            ],
            "approvals": [
                {
                    "approval_id": ap.id,
                    "sequence_number": ap.sequence_number,
                    "is_mandatory": ap.is_mandatory,
                    "reconsideration_requested": ap.reconsideration_requested,
                    "status": ap.status.value,
                    "reviewed_at": ap.reviewed_at,
                    "message": ap.message,
                    "user": {
                        "id": ap.user.id,
                        "name": ap.user.name,
                        "email": ap.user.email,
                        "nickname": ap.user.nickname,
                        "num_document": ap.user.num_document
                    }
                }
                for ap in r.approvals
            ]
        })

    return result


@router.get("/users/completed_forms_with_responses")
def get_completed_forms_with_responses(
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    """
    Retorna los formularios completados por el usuario autenticado junto con 
    todas sus respuestas y aprobaciones en una sola llamada.
    
    Combina la información de:
    - Lista de formularios completados
    - Resumen de todas las respuestas de cada formulario con sus aprobaciones

    - **Autenticación requerida**
    - **Código 200**: Lista de formularios con sus respuestas y aprobaciones
    - **Código 403**: Usuario no autenticado o sin permisos
    - **Código 404**: No se encontraron formularios completados
    """
    if not current_user:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to access completed forms",
        )

    completed_forms_data = fetch_completed_forms_with_all_responses(db, current_user.id)
    
    if not completed_forms_data:
        raise HTTPException(
            status_code=404, 
            detail="No completed forms found for this user"
        )
    
    return completed_forms_data

# Endpoint para las estadisticas resumidas de las respuestas
@router.get("/responses/summary")
def get_responses_summary(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Obtiene un resumen de las respuestas del usuario autenticado para un formulario específico.
    Solo devuelve campos básicos: información del formulario, estado de aprobación y datos mínimos.
    
    Ideal para listados rápidos donde no se necesitan las respuestas completas a las preguntas.

    Args:
        form_id (int): ID del formulario del cual se desean obtener las respuestas.
        db (Session): Sesión activa de la base de datos.
        current_user (User): Usuario autenticado.

    Returns:
        List[dict]: Lista resumida de respuestas con información básica de aprobación.

    Raises:
        HTTPException: 403 si no hay un usuario autenticado.
        HTTPException: 404 si no se encuentran respuestas.
    """
    if not current_user:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to access responses",
        )

    # Query optimizado: solo traemos lo necesario
    stmt = (
        select(Response)
        .where(Response.form_id == form_id, Response.user_id == current_user.id)
        .options(
            joinedload(Response.form).load_only(
                Form.id,
                Form.title,
                Form.description
            ),
            joinedload(Response.approvals).joinedload(ResponseApproval.user).load_only(
                User.id,
                User.name,
                User.email,
                User.nickname,
                User.num_document
            )
        )
    )

    responses = db.execute(stmt).unique().scalars().all()

    if not responses:
        raise HTTPException(status_code=404, detail="No se encontraron respuestas")

    result = []
    for r in responses:
        approval_result = get_response_approval_status(r.approvals)

        result.append({
            "form_id": r.form.id,
            "form_title": r.form.title,
            "form_description": r.form.description,
            "response_id": r.id,
            "submitted_at": r.submitted_at,
            "approvals": [
                {
                    "approval_id": ap.id,
                    "sequence_number": ap.sequence_number,
                    "is_mandatory": ap.is_mandatory,
                    "user": {
                        "id": ap.user.id,
                        "name": ap.user.name

                    }
                }
                for ap in r.approvals
            ]
        })

    return result

@router.get("/all/list", response_model=List[dict])
def get_forms_endpoint(db: Session = Depends(get_db)):
    """
    Retorna una lista de todos los formularios registrados en la base de datos.

    - **Retorna**: Lista de diccionarios con la información de cada formulario.
    - **Código 200**: Éxito, formularios encontrados.
    - **Código 404**: No se encontraron formularios.
    - **Código 500**: Error interno del servidor.
    """
    try:
        forms = get_all_forms(db)
        if not forms:
            raise HTTPException(status_code=404, detail="No se encontraron formularios")
        return forms
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@router.get("/all/list/paginated")
def get_forms_paginated_endpoint(
    page: int = 1,
    page_size: int = 30,
    db: Session = Depends(get_db)
):
    """
    Retorna una lista paginada de todos los formularios registrados en la base de datos.

    - **page**: Número de página (por defecto 1)
    - **page_size**: Cantidad de registros por página (por defecto 30, máximo 100)
    - **Retorna**: Diccionario con items paginados y metadata.
    - **Código 200**: Éxito, formularios encontrados.
    - **Código 404**: No se encontraron formularios.
    - **Código 500**: Error interno del servidor.
    """
    try:
        # Validar page_size máximo
        if page_size > 100:
            page_size = 100
        
        forms_data = get_all_forms_paginated(db, page, page_size)
        
        if not forms_data["items"]:
            raise HTTPException(status_code=404, detail="No se encontraron formularios")
        
        return forms_data
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/users/form_by_user")
def get_user_forms(
    page: int = 1,  # Número de página (empieza en 1)
    page_size: int = 30,  # Cantidad de registros por página
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    """
    Retorna los formularios asignados al usuario autenticado con paginación.

    - **page**: Número de página (por defecto 1)
    - **page_size**: Cantidad de registros por página (por defecto 30, máximo 100)
    - **Requiere autenticación.**
    - **Código 200**: Lista de formularios paginados.
    - **Código 403**: Usuario sin permisos.
    - **Código 404**: No se encontraron formularios.
    - **Código 500**: Error interno del servidor.
    """
    try:
        if current_user is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="User does not have permission to get all questions"
            )
        
        # Validar page_size máximo
        if page_size > 100:
            page_size = 100
        
        # Obtener formularios paginados
        forms_data = get_forms_by_user(db, current_user.id, page, page_size)
        
        if not forms_data["items"]:
            raise HTTPException(
                status_code=404, 
                detail="No se encontraron formularios para este usuario"
            )
        
        return forms_data
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/users/form_by_user/summary")
def get_user_forms_summary(
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    """
    Retorna un resumen de los formularios asignados al usuario autenticado.
    Solo devuelve campos básicos: id, title, description, created_at, user_id
    
    Ideal para listados rápidos donde no se necesita información completa.
    
    - **Requiere autenticación.**
    - **Código 200**: Lista resumida de formularios.
    - **Código 403**: Usuario sin permisos.
    - **Código 404**: No se encontraron formularios.
    - **Código 500**: Error interno del servidor.
    """
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to get forms"
        )
    
    try:
        forms = get_forms_by_user_summary(db, current_user.id)
        if not forms:
            raise HTTPException(
                status_code=404, 
                detail="No se encontraron formularios para este usuario"
            )
        return forms
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/users/completed_forms")
def get_completed_forms_for_user(
    page: int = 1,
    page_size: int = 30,
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    """
    Retorna los formularios que han sido completados por el usuario autenticado con paginación.

    - **page**: Número de página (por defecto 1)
    - **page_size**: Cantidad de registros por página (por defecto 30, máximo 100)
    - **Autenticación requerida**
    - **Código 200**: Lista paginada de formularios completados
    - **Código 403**: Usuario no autenticado o sin permisos
    - **Código 404**: No se encontraron formularios completados
    """
    if not current_user:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to access completed forms",
        )

    # Validar page_size máximo
    if page_size > 100:
        page_size = 100
    
    # Validar que page sea mayor a 0
    if page < 1:
        page = 1

    completed_forms_data = fetch_completed_forms_by_user(db, current_user.id, page, page_size)
    
    if not completed_forms_data["items"]:
        raise HTTPException(status_code=404, detail="No completed forms found for this user")
    
    return completed_forms_data


@router.get("/{form_id}/questions_associated_and_unassociated")
def get_form_questions(form_id: int, db: Session = Depends(get_db),current_user: User = Depends(get_current_user)):
    """
    Endpoint para obtener las preguntas asociadas y no asociadas a un formulario dado su ID.

    Solo los usuarios con tipo 'creator' o 'admin' pueden acceder a esta funcionalidad.

    Args:
        form_id (int): ID del formulario.
        db (Session): Sesión de base de datos proporcionada por FastAPI.
        current_user (User): Usuario autenticado obtenido a través del sistema de dependencias.

    Returns:
        dict: Diccionario con dos listas: 'associated_questions' y 'unassociated_questions'.

    Raises:
        HTTPException: Si el usuario no tiene permisos o si el formulario no existe.
    """
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )

    return fetch_form_questions(form_id, db)



@router.post("/{form_id}/questions/{question_id}")
def add_question_to_form(form_id: int, question_id: int, db: Session = Depends(get_db),current_user: User = Depends(get_current_user)):
    """
    Asocia una pregunta existente a un formulario específico.
    Solo los usuarios con rol 'creator' o 'admin' pueden realizar esta acción.
    Args:
        form_id (int): ID del formulario.
        question_id (int): ID de la pregunta a asociar.
        db (Session): Sesión de base de datos proporcionada por FastAPI.
        current_user (User): Usuario autenticado.

    Returns:
        dict: Mensaje de éxito y ID de la relación creada.

    """
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )
        
    return link_question_to_form(form_id, question_id, db)

@router.get("/{form_id}/users_associated_and_unassociated")
def get_form_users(form_id: int, db: Session = Depends(get_db),current_user: User = Depends(get_current_user) ):
    """
    Obtiene los usuarios asociados y no asociados como moderadores a un formulario.

    Solo los usuarios con rol 'creator' o 'admin' pueden acceder.

    Args:
        form_id (int): ID del formulario.
        db (Session): Sesión de base de datos proporcionada por FastAPI.
        current_user (User): Usuario autenticado.

    Returns:
        dict: Diccionario con dos listas: 'associated_users' y 'unassociated_users'.

    Raises:
        HTTPException:
            - 403: Si el usuario no tiene permisos.
            - 404: Si el formulario no existe.
    """
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )
    return fetch_form_users(form_id, db)

@router.post("/{form_id}/form_moderators/{user_id}")
def add_user_to_form_schedule(form_id: int, user_id: int, db: Session = Depends(get_db),current_user: User = Depends(get_current_user)):
    """
    Asocia un usuario como moderador de un formulario.
    Solo los usuarios con rol 'creator' o 'admin' pueden acceder.
    Args:
        form_id (int): ID del formulario.
        user_id (int): ID del usuario a asociar como moderador.
        db (Session): Sesión de base de datos proporcionada por FastAPI.
        current_user (USer): Usuario autenticado.

    Returns:
        dict: Mensaje de éxito con el ID de la relación creada.
    Raises:
        HTTPException:
            - 403: Si el usuario no tiene permisos.
            - 404: Si el formulario o el usuario no existen.
            - 400: Si ya existe la relación.
    """ 
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )
    return link_moderator_to_form(form_id, user_id, db)


@router.delete("/{form_id}/questions/{question_id}/delete")
def delete_question_from_form(form_id: int, question_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Elimina la relación entre una pregunta y un formulario.

    Solo accesible para usuarios con rol 'creator' o 'admin'.

    Args:
        form_id (int): ID del formulario.
        question_id (int): ID de la pregunta a desvincular.
        db (Session): Sesión de base de datos proporcionada por FastAPI.
        current_user (User): Usuario autenticado.

    Returns:
        dict: Mensaje de éxito.
    """ 
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )
    return remove_question_from_form(form_id, question_id, db)

@router.delete("/{form_id}/moderators/{user_id}/delete")
def delete_moderator_from_form(form_id: int, user_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Elimina la relación entre una pregunta y un formulario.
    Solo accesible para usuarios con rol 'creator' o 'admin'.
    Args:
        form_id (int): ID del formulario.
        question_id (int): ID de la pregunta a desvincular.
        db (Session): Sesión de base de datos proporcionada por FastAPI.
        current_user (User): Usuario autenticado.

    Returns:
        dict: Mensaje de éxito.
    """
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )
    return remove_moderator_from_form(form_id, user_id, db)

@router.post("/form-answers/")
def create_form_answer(payload: FormAnswerCreate, db: Session = Depends(get_db)):
    """
    Crea una nueva relación entre un formulario y una pregunta en la tabla FormAnswer.
    Esta relación se usa para definir si una pregunta está repetida en un formulario.
    Args:
        payload (FormAnswerCreate): Datos requeridos para crear la relación, incluyendo:
            - form_id (int): ID del formulario.
            - question_id (int): ID de la pregunta.
            - is_repeated (bool): Indicador si la pregunta se repite.

        db (Session): Sesión de base de datos proporcionada por FastAPI.

    Returns:
        dict: Mensaje de confirmación y datos de la relación creada.
    """
    form_answer = FormAnswer(
        form_id=payload.form_id,
        question_id=payload.question_id,
        is_repeated=payload.is_repeated
    )
    db.add(form_answer)
    db.commit()
    db.refresh(form_answer)
    return {
        "message": "FormAnswer created successfully",
        "data": {
            "id": form_answer.id,
            "form_id": form_answer.form_id,
            "question_id": form_answer.question_id,
            "is_repeated": form_answer.is_repeated
        }
    }

@router.post("/forms-by-answers/", response_model=List[FormResponse])
def get_forms_by_answers(
    answer_ids: List[int], 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """
    Endpoint que obtiene los formularios asociados a respuestas y verifica si el usuario es moderador de ellos.
    """
    forms = get_moderated_forms_by_answers(answer_ids, current_user.id, db)
    return forms

@router.get("/{form_id}/questions-answers/excel")
def download_questions_answers_excel(form_id: int, db: Session = Depends(get_db)):
    """
    Genera un archivo Excel con las preguntas y respuestas de un formulario específico.
    """
    data = get_questions_and_answers_by_form_id(db, form_id)
    if not data:
        raise HTTPException(status_code=404, detail="Formulario no encontrado")

    df = pd.DataFrame(data["data"])
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name="Respuestas")
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Formulario_{form_id}_respuestas.xlsx"}
    )


@router.get("/{form_id}/answers/excel/all-users")
def download_questions_answers_excel_all_users(form_id: int, db: Session = Depends(get_db)):
    """
    Genera un archivo Excel con las preguntas y respuestas de un formulario específico.
    Ruta alternativa para uso en emails automáticos.
    """
    data = get_questions_and_answers_by_form_id(db, form_id)
    if not data:
        raise HTTPException(status_code=404, detail="Formulario no encontrado")

    df = pd.DataFrame(data["data"])
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name="Respuestas")
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Formulario_{form_id}_respuestas.xlsx"}
    )


@router.get("/{form_id}/questions-answers/excel/user")
def download_user_responses_excel(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Exporta las respuestas del usuario a Excel.
    Si el formulario tiene form_design, replica el diseño visual completo.
    Si no, hace fallback al formato simple con pandas.

    v3.0: Reconstruye repeated_id desde form_design + FormAnswer.is_repeated
    """
    from sqlalchemy.orm import joinedload
    import json

    # 1. Obtener el formulario
    form = db.query(Form).filter(Form.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Formulario no encontrado")

    # 2. Verificar si tiene form_design
    form_design = form.form_design

    if isinstance(form_design, str):
        try:
            form_design = json.loads(form_design)
        except json.JSONDecodeError:
            form_design = None

    # 3. Si tiene form_design → generar Excel con diseño
    if form_design and isinstance(form_design, list) and len(form_design) > 0:

        responses = (
            db.query(Response)
            .filter(Response.form_id == form_id, Response.user_id == current_user.id)
            .order_by(Response.submitted_at.desc())
            .all()
        )

        if not responses:
            raise HTTPException(
                status_code=404,
                detail="No se encontraron respuestas para este usuario"
            )

        latest = responses[0]

        answers_orm = (
            db.query(Answer)
            .options(joinedload(Answer.question))
            .filter(Answer.response_id == latest.id)
            .all()
        )

        # ✅ Serializar CON reconstrucción de repeated_id
        answers = _serialize_answers(answers_orm, db, form_id, form_design)

        style_config = _extract_style_config(form_design)

        output = generate_form_excel(
            form_design=form_design,
            answers=answers,
            style_config=style_config,
            form_title=form.title,
            response_id=latest.id,
        )

        filename = f"Formato_{form.title.replace(' ', '_')}_{form_id}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # 4. Fallback: formato simple con pandas
    data = get_questions_and_answers_by_form_id_and_user(db, form_id, current_user.id)
    if not data or not data["data"]:
        raise HTTPException(
            status_code=404,
            detail="No se encontraron respuestas para este usuario en el formulario"
        )

    df = pd.DataFrame(data["data"])
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name="Respuestas del Usuario")
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=Respuestas_usuario_{current_user.id}_formulario_{form_id}.xlsx"
        }
    )
                                                                                                                                                                                                                                     
                      
@router.get("/{form_id}/responses_data_forms")
def get_form_responses(form_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Obtiene todas las respuestas asociadas a un formulario específico.

    - **form_id**: ID del formulario a consultar.
    - **current_user**: Usuario autenticado que realiza la solicitud.
    - **db**: Sesión de base de datos.

    Retorna los datos del formulario incluyendo:
    - Información del formulario.
    - Respuestas enviadas por los usuarios.
    - Información del usuario que respondió.
    - Preguntas y respuestas (texto y archivos si aplica).

    Requiere que el usuario tenga tipo `creator` o `admin`.
    """
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )
    data = get_form_responses_data(form_id, db)
    if not data:
        raise HTTPException(status_code=404, detail="Formulario no encontrado")
    return data                                                                              

@router.get("/{user_id}/responses_data_users")
def get_user_responses(user_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Obtiene todas las respuestas asociadas a un usuario específico.

    - **user_id**: ID del usuario a consultar.
    - **current_user**: Usuario autenticado que realiza la solicitud.
    - **db**: Sesión de base de datos.

    Requiere permisos de tipo `creator` o `admin`.

    Retorna un diccionario con:
    - Información del usuario.
    - Todas las respuestas que ha enviado.
    - Información del formulario relacionado.
    - Preguntas y sus respectivas respuestas.
    """
    
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )
    data = get_user_responses_data(user_id, db)
    if not data:
        raise HTTPException(status_code=404, detail="User or responses not found")
    return data

@router.get("/{form_id}/questions-answers/excel/user/{id_user}")
def download_user_responses_excel(
    form_id: int,
    id_user: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Descarga en formato Excel las respuestas de un usuario específico.
    Si tiene form_design → replica el diseño visual completo.
    Si no → fallback al formato simple con pandas.
    """
    from sqlalchemy.orm import joinedload
    from copy import copy
    from openpyxl import Workbook, load_workbook


    # Validar permisos
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )

    # 1. Obtener el formulario
    form = db.query(Form).filter(Form.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Formulario no encontrado")

    # 2. Obtener form_design
    form_design = form.form_design

    if isinstance(form_design, str):
        import json
        try:
            form_design = json.loads(form_design)
        except json.JSONDecodeError:
            form_design = None

    # 3. Si tiene form_design → generar Excel con diseño
    if form_design and isinstance(form_design, list) and len(form_design) > 0:

        # Consultar respuestas del usuario específico
        all_responses = (
            db.query(Response)
            .filter(Response.form_id == form_id, Response.user_id == id_user)
            .order_by(Response.submitted_at.desc())
            .all()
        )

        if not all_responses:
            raise HTTPException(
                status_code=404,
                detail="No se encontraron respuestas para este usuario"
            )

        # Extraer style_config
        style_config = None
        for item in form_design:
            props = item.get("props") or {}
            if props.get("styleConfig"):
                style_config = props["styleConfig"]
                break
            if item.get("headerTable") and not item.get("type"):
                style_config = item
                break

        # Nombre del usuario
        target_user = db.query(User).filter(User.id == id_user).first()
        user_name = target_user.name if target_user else f"Usuario_{id_user}"
        safe_name = "".join(c for c in user_name if c.isalnum() or c in " _-")[:15]

        # Si solo hay 1 respuesta → una sola hoja
        if len(all_responses) == 1:
            resp = all_responses[0]
            answers_orm = (
                db.query(Answer)
                .options(joinedload(Answer.question))
                .filter(Answer.response_id == resp.id)
                .all()
            )

            answers = []
            for ans in answers_orm:
                answers.append({
                    "id_answer":               ans.id,
                    "question_id":             ans.question_id,
                    "question_text":           ans.question.question_text if ans.question else "",
                    "question_type":           ans.question.question_type.value if (
                                                   ans.question and ans.question.question_type
                                               ) else "text",
                    "answer_text":             ans.answer_text or "",
                    "file_path":               ans.file_path or "",
                    "repeated_id":             getattr(ans, "repeated_id", None),
                    "form_design_element_id":  getattr(ans, "form_design_element_id", None),
                })

            output = generate_form_excel(
                form_design=form_design,
                answers=answers,
                style_config=style_config,
                form_title=form.title,
                response_id=resp.id,
            )

            filename = f"Respuestas_{safe_name}_{form.title.replace(' ', '_')}_{form_id}.xlsx"

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )

        # Si hay múltiples respuestas → una hoja por cada una
        wb_final = Workbook()
        wb_final.remove(wb_final.active)

        for resp in all_responses:
            answers_orm = (
                db.query(Answer)
                .options(joinedload(Answer.question))
                .filter(Answer.response_id == resp.id)
                .all()
            )

            answers = []
            for ans in answers_orm:
                answers.append({
                    "id_answer":               ans.id,
                    "question_id":             ans.question_id,
                    "question_text":           ans.question.question_text if ans.question else "",
                    "question_type":           ans.question.question_type.value if (
                                                   ans.question and ans.question.question_type
                                               ) else "text",
                    "answer_text":             ans.answer_text or "",
                    "file_path":               ans.file_path or "",
                    "repeated_id":             getattr(ans, "repeated_id", None),
                    "form_design_element_id":  getattr(ans, "form_design_element_id", None),
                })

            single_output = generate_form_excel(
                form_design=form_design,
                answers=answers,
                style_config=style_config,
                form_title=form.title,
                response_id=resp.id,
            )

            temp_wb = load_workbook(single_output)
            temp_ws = temp_wb.active

            sheet_name = f"R{resp.id}"[:31]
            existing_names = wb_final.sheetnames
            if sheet_name in existing_names:
                counter = 2
                while f"{sheet_name[:28]}_{counter}" in existing_names:
                    counter += 1
                sheet_name = f"{sheet_name[:28]}_{counter}"

            new_ws = wb_final.create_sheet(title=sheet_name)

            # Copiar celdas con copy() para evitar StyleProxy error
            for row in temp_ws.iter_rows():
                for cell in row:
                    new_cell = new_ws.cell(
                        row=cell.row,
                        column=cell.column,
                        value=cell.value
                    )
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.alignment = copy(cell.alignment)
                        new_cell.border = copy(cell.border)
                        new_cell.number_format = cell.number_format

            for col_letter, dim in temp_ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = dim.width
            for row_num, dim in temp_ws.row_dimensions.items():
                new_ws.row_dimensions[row_num].height = dim.height
            for merge_range in temp_ws.merged_cells.ranges:
                new_ws.merge_cells(str(merge_range))

        output = BytesIO()
        wb_final.save(output)
        output.seek(0)

        filename = f"Respuestas_{safe_name}_{form.title.replace(' ', '_')}_{form_id}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # 4. Fallback: formato simple con pandas
    data = get_questions_and_answers_by_form_id_and_user(db, form_id, id_user)

    if not data or not data["data"]:
        raise HTTPException(
            status_code=404,
            detail="No se encontraron respuestas para este usuario en el formulario"
        )

    df = pd.DataFrame(data["data"])
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name="Respuestas del Usuario")
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=Respuestas_usuario_{id_user}_formulario_{form_id}.xlsx"
        }
    )
@router.get("/form-schedules_table/", response_model=List[FormScheduleOut], )
def get_form_schedules(form_id: int, user_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Devuelve las programaciones (`FormSchedule`) asociadas a un formulario y usuario específicos.

    - **form_id**: ID del formulario del cual se desea obtener las programaciones.
    - **user_id**: ID del usuario al que están asociadas las programaciones.
    - **current_user**: Usuario autenticado (verificado).
    - **db**: Sesión de base de datos proporcionada por la dependencia `get_db`.

    Retorna:
    - Lista de objetos `FormScheduleOut` que contienen la información de las programaciones.

    Errores posibles:
    - **403**: Si no hay usuario autenticado.
    - **404**: Si no se encuentran programaciones para ese formulario y usuario.
    """ 
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to get all questions"
        )
    else:
        schedules = db.query(FormSchedule).filter(
            FormSchedule.form_id == form_id,
            FormSchedule.user_id == user_id
        ).all()

        if not schedules:
            raise HTTPException(status_code=404, detail="No se encontraron programaciones para ese formulario y usuario.")

        return schedules



@router.get("/{form_id}/questions-answers/excel/all-users")
def download_all_user_responses_excel(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Descarga TODAS las respuestas de TODOS los usuarios para un formulario.
    Si tiene form_design → genera un Excel con una hoja por respuesta, replicando el diseño.
    Si no → fallback al formato simple.
    """
    from sqlalchemy.orm import joinedload
    from openpyxl import Workbook, load_workbook
    from copy import copy

    # 1. Obtener el formulario
    form = db.query(Form).filter(Form.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Formulario no encontrado")

    # 2. Obtener form_design
    form_design = form.form_design

    if isinstance(form_design, str):
        import json
        try:
            form_design = json.loads(form_design)
        except json.JSONDecodeError:
            form_design = None

    # 3. Si tiene form_design → generar Excel con diseño
    if form_design and isinstance(form_design, list) and len(form_design) > 0:

        all_responses = (
            db.query(Response)
            .filter(Response.form_id == form_id)
            .order_by(Response.submitted_at.desc())
            .all()
        )

        if not all_responses:
            raise HTTPException(
                status_code=404,
                detail="No se encontraron respuestas para este formulario"
            )

        style_config = _extract_style_config(form_design)

        # Crear workbook final
        wb_final = Workbook()
        wb_final.remove(wb_final.active)

        for resp in all_responses:
            # Obtener answers
            answers_orm = (
                db.query(Answer)
                .options(joinedload(Answer.question))
                .filter(Answer.response_id == resp.id)
                .all()
            )

            # Nombre del usuario
            user = db.query(User).filter(User.id == resp.user_id).first()
            user_name = user.name if user else f"Usuario_{resp.user_id}"
            safe_name = "".join(c for c in user_name if c.isalnum() or c in " _-")[:15]

            # ✅ CORREGIDO: Usar _serialize_answers para reconstruir repeated_id
            answers = _serialize_answers(answers_orm, db, form_id, form_design)

            # Generar Excel individual
            single_output = generate_form_excel(
                form_design=form_design,
                answers=answers,
                style_config=style_config,
                form_title=form.title,
                response_id=resp.id,
            )

            # Cargar el workbook temporal
            temp_wb = load_workbook(single_output)
            temp_ws = temp_wb.active

            # Nombre de hoja único
            sheet_name = f"R{resp.id}_{safe_name}"[:31]
            existing_names = wb_final.sheetnames
            if sheet_name in existing_names:
                counter = 2
                while f"{sheet_name[:28]}_{counter}" in existing_names:
                    counter += 1
                sheet_name = f"{sheet_name[:28]}_{counter}"

            new_ws = wb_final.create_sheet(title=sheet_name)

            # Copiar celdas con estilos
            for row in temp_ws.iter_rows():
                for cell in row:
                    new_cell = new_ws.cell(
                        row=cell.row,
                        column=cell.column,
                        value=cell.value
                    )
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.alignment = copy(cell.alignment)
                        new_cell.border = copy(cell.border)
                        new_cell.number_format = cell.number_format

            # Copiar anchos de columna
            for col_letter, dim in temp_ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = dim.width

            # Copiar alturas de fila
            for row_num, dim in temp_ws.row_dimensions.items():
                new_ws.row_dimensions[row_num].height = dim.height

            # Copiar merges
            for merge_range in temp_ws.merged_cells.ranges:
                new_ws.merge_cells(str(merge_range))

        # Guardar
        output = BytesIO()
        wb_final.save(output)
        output.seek(0)

        filename = f"Todas_Respuestas_{form.title.replace(' ', '_')}_{form_id}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # 4. Fallback: formato simple
    data = get_all_user_responses_by_form_id_improved(db, form_id)

    if not data or not data["data"]:
        raise HTTPException(
            status_code=404,
            detail="No se encontraron respuestas para este formulario"
        )

    output = generate_excel_with_repeaters(data)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=Respuestas_formulario_{form_id}_usuarios.xlsx"
        }
    )
@router.get("/users/unanswered_forms",
    summary="Obtener formularios no respondidos",
    description="Retorna los formularios asignados al usuario autenticado que aún no han sido respondidos."
)
def get_unanswered_forms(
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    try:
        if current_user is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="User does not have permission to get forms"
            )

        forms = get_unanswered_forms_by_user(db, current_user.id)

        if not forms:
            raise HTTPException(status_code=404, detail="No hay formularios sin responder para este usuario")

        return forms

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/responses/by-user/")
def get_responses_by_user_and_form(
    form_id: int,
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtiene todas las Responses junto con sus Answers basado en form_id y user_id específicos.
    Requiere permisos de administrador o autorización adecuada.
    """
    # Verifica permisos si es necesario, por ejemplo, que sea admin
    if current_user is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="User does not have permission to get forms"
            )


    stmt = (
        select(Response)
        .where(Response.form_id == form_id, Response.user_id == user_id)
        .options(
            joinedload(Response.answers).joinedload(Answer.question)
        )
    )

    results = db.execute(stmt).unique().scalars().all()

    if not results:
        raise HTTPException(status_code=404, detail="No se encontraron respuestas")

    return results

@router.post("/create/response_approval_endpoint", status_code=status.HTTP_201_CREATED)
def create_response_approval_endpoint(
    data: ResponseApprovalCreate,
    db: Session = Depends(get_db)
):
    """
    Crea un nuevo registro de aprobación de respuesta.

    Este endpoint recibe los datos necesarios para crear un registro en la tabla `ResponseApproval`
    y lo almacena en la base de datos.

    Parámetros:
    ----------
    data : ResponseApprovalCreate
        Objeto que contiene los datos requeridos para crear la aprobación.
    db : Session
        Sesión activa de la base de datos (inyectada automáticamente por FastAPI).

    Retorna:
    -------
    ResponseApproval
        Objeto creado de tipo ResponseApproval.

    Excepciones:
    -----------
    HTTPException (400):
        Si ocurre un error durante la creación del registro, se retorna una excepción con el mensaje de error.
    """
    try:
        return create_response_approval(db, data)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    

@router.get("/user/assigned-forms-with-responses")
def get_forms_to_approve( db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Obtiene los formularios asignados al usuario actual que tienen respuestas pendientes de aprobación.

    Este endpoint retorna una lista de formularios con sus respuestas correspondientes que requieren 
    la aprobación del usuario autenticado, respetando el orden de secuencia y verificando si es su turno 
    de aprobar según las reglas definidas.

    Parámetros:
    ----------
    db : Session
        Sesión activa de la base de datos (inyectada por FastAPI).
    current_user : User
        Usuario autenticado (inyectado por el sistema de autenticación).

    Retorna:
    -------
    List[Dict]
        Lista de formularios con información detallada de las respuestas, 
        aprobaciones y el estado de cada aprobador.
    """
    return get_forms_pending_approval_for_user(current_user.id, db)

@router.post("/create_notification")
def create_notification(notification: NotificationCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Crea una notificación para eventos de aprobación de formularios.

    Este endpoint permite registrar una notificación que se activará cuando se cumpla la condición 
    especificada (por ejemplo, una nueva respuesta o aprobación).

    Parámetros:
    ----------
    notification : NotificationCreate
        Objeto con los datos necesarios para crear la notificación. Contiene:
        - form_id: ID del formulario al cual aplica la notificación.
        - user_id: ID del usuario que debe recibir la notificación.
        - notify_on: Evento que dispara la notificación (por ejemplo: "on_submit", "on_approval").

    db : Session
        Sesión de base de datos (inyectada por FastAPI).
    
    current_user : User
        Usuario autenticado. Se utiliza para validar si tiene permisos.

    Validaciones:
    ------------
    - Si el usuario no está autenticado, se retorna un error 403.
    - Si ya existe una notificación con los mismos datos (formulario, usuario y tipo de evento),
      se retorna un error 400 para evitar duplicados.

    Retorna:
    -------
    dict
        Mensaje de éxito con el ID de la notificación creada:
        {
            "message": "Notificación creada correctamente",
            "id": <id_notificación>
        }

    Errores:
    -------
    - 403 FORBIDDEN: Si el usuario no tiene permisos.
    - 400 BAD REQUEST: Si la notificación ya existe.
    """
    
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to get forms"
        )
    # Verifica si ya existe una notificación similar (opcional)
    existing = db.query(FormApprovalNotification).filter_by(
        form_id=notification.form_id,
        user_id=notification.user_id,
        notify_on=notification.notify_on
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Esta notificación ya existe.")

    new_notification = FormApprovalNotification(
        form_id=notification.form_id,
        user_id=notification.user_id,
        notify_on=notification.notify_on
    )
    db.add(new_notification)
    db.commit()
    db.refresh(new_notification)
    return {"message": "Notificación creada correctamente", "id": new_notification.id}

 
@router.get("/form-details/{form_id}")
def get_form_details(form_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    
    """
    Retorna los detalles completos de un formulario, incluidas preguntas, respuestas y estado de aprobación.

    Este endpoint recupera:
    - Información básica del formulario (ID, título, descripción).
    - Preguntas asociadas al formulario.
    - Respuestas completas con información del usuario que respondió.
    - Historial de respuestas (si las respuestas fueron editadas).
    - Estado de aprobación por respuesta.

    Requiere autenticación.

    Parámetros:
    ----------
    form_id : int
        ID del formulario a consultar.

    db : Session
        Sesión activa de la base de datos.

    current_user : User
        Usuario autenticado (obligatorio).

    Retorna:
    -------
    dict
        Objeto con la estructura del formulario, preguntas, respuestas, aprobaciones e historial.

    Errores:
    -------
    - 403 FORBIDDEN: Si el usuario no está autenticado.
    - 404 NOT FOUND: Si el formulario no existe.
    """
    
    if current_user == None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to get options"
            )
    else: 
        result = get_form_with_full_responses(form_id, db)

        if not result:
            raise HTTPException(status_code=404, detail="Formulario no encontrado")

        return result
    
    
@router.put("/{form_id}/design", status_code=200)
def update_form_design(
    form_id: int,
    payload: FormDesignUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Actualiza el diseño visual del formulario especificado.

    Este endpoint permite modificar el diseño personalizado de un formulario,
    reemplazando completamente su estructura `form_design` con la nueva proporcionada.

    Parámetros:
    -----------
    form_id : int
        ID del formulario a actualizar.

    payload : FormDesignUpdate
        Datos del nuevo diseño a aplicar. Contiene:
        - `form_design`: una lista de objetos/diccionarios con la estructura del diseño.

    db : Session
        Sesión activa de la base de datos.

    current_user : User
        Usuario autenticado que realiza la acción.

    Retorna:
    --------
    list[dict]
        Lista con un mensaje de confirmación y el ID del formulario actualizado:
        ```json
        [
            {
                "message": "Form design updated successfully",
                "form_id": 123
            }
        ]
        ```

    Errores:
    --------
    - 403 FORBIDDEN: Si el usuario no está autenticado.
    - 404 NOT FOUND: Si el formulario no existe.
    """
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to update form design"
        )

    updated_form = update_form_design_service(db, form_id, payload.form_design)
    
    # 🔥 INVALIDAR CACHÉ DE REDIS
    invalidate_form_cache(form_id)
    
    return [{
        "message": "Form design updated successfully",
        "form_id": updated_form.id
    }]


@router.get("/{form_id}/notifications", response_model=NotificationsByFormResponse_schema)
def get_notifications_by_form(form_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # Llamamos a la función para obtener las notificaciones
    """
    Obtiene todas las notificaciones configuradas para un formulario específico.

    Este endpoint devuelve una lista de notificaciones que han sido definidas para
    el formulario identificado por `form_id`. Cada notificación incluye información
    del tipo de notificación y del usuario al que se notificará.

    Solo usuarios autenticados pueden acceder a esta información.

    Parámetros:
    -----------
    form_id : int
        ID del formulario para el cual se desean obtener las notificaciones.

    db : Session
        Sesión activa de la base de datos proporcionada por FastAPI.

    current_user : User
        Usuario autenticado que realiza la solicitud.

    Retorna:
    --------
    NotificationsByFormResponse_schema
        Objeto con el ID del formulario y la lista de notificaciones asociadas.

    Lanza:
    ------
    HTTPException 403:
        Si el usuario no está autenticado o no tiene permisos para ver las notificaciones.
    
    HTTPException 404:
        Si el formulario no existe (lanzado desde la función auxiliar `get_notifications_for_form`).
    """
    if current_user == None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to get options"
            )
    else: 
        notifications = get_notifications_for_form(form_id, db)

        # Preparamos la respuesta con las notificaciones
        return NotificationsByFormResponse_schema(
            form_id=form_id,
            notifications=notifications
        )
        
@router.put("/notifications/update-status/{notification_id}", response_model=UpdateNotifyOnSchema)
def update_notify_status_endpoint(notification_id: int, request: UpdateNotifyOnSchema, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Actualiza el tipo de notificación ('notify_on') de una notificación específica.

    Este endpoint permite cambiar el tipo de evento que desencadenará la notificación,
    como "cada_aprobacion" o "aprobacion_final", para una notificación ya existente
    asociada a un formulario.

    Requiere autenticación.

    Parámetros:
    -----------
    notification_id : int
        ID de la notificación a actualizar.

    request : UpdateNotifyOnSchema
        Objeto que contiene el nuevo valor del campo `notify_on`.

    db : Session
        Sesión activa de base de datos proporcionada por FastAPI.

    current_user : User
        Usuario autenticado que realiza la solicitud.

    Retorna:
    --------
    UpdateNotifyOnSchema:
        Datos actualizados de la notificación.

    Lanza:
    ------
    HTTPException 403:
        Si el usuario no está autenticado o no tiene permisos.

    HTTPException 400:
        Si el valor de `notify_on` no es válido.

    HTTPException 404:
        Si la notificación no existe.
    """
    if current_user == None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to get options"
            )
    else: 
        return update_notification_status(notification_id, request.notify_on, db)
    

@router.delete("/notifications/{notification_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_notification(notification_id: int, db: Session = Depends(get_db),  current_user: User = Depends(get_current_user)):
    """
    Elimina una notificación específica de un formulario.

    Este endpoint permite eliminar una notificación creada previamente que
    está asociada a un formulario de aprobación. Solo usuarios autenticados
    con permisos pueden realizar esta acción.

    Parámetros:
    -----------
    notification_id : int
        ID de la notificación que se desea eliminar.

    db : Session
        Sesión activa de la base de datos proporcionada por FastAPI.

    current_user : User
        Usuario autenticado que realiza la solicitud.

    Retorna:
    --------
    dict:
        Mensaje de confirmación de eliminación (aunque el status sea 204).

    Lanza:
    ------
    HTTPException 403:
        Si el usuario no tiene permisos o no está autenticado.

    HTTPException 404:
        Si la notificación no existe.
    """
    if current_user == None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to get options"
            )
    else: 
        notification = db.query(FormApprovalNotification).filter(FormApprovalNotification.id == notification_id).first()

        if not notification:
            # Si no existe, lanzar un error 404
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Notificación no encontrada")
        
        # Eliminar la notificación
        db.delete(notification)
        db.commit()

        return {"message": "Notificación eliminada correctamente"}
    
@router.delete("/notifications/bulk/delete", status_code=status.HTTP_204_NO_CONTENT)
def delete_notifications_bulk(
    notification_ids: List[int],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Elimina múltiples notificaciones de una sola vez.
    
    Este endpoint permite eliminar varios notificadores de forma masiva,
    mejorando el rendimiento cuando hay muchos registros.
    
    Parámetros:
    -----------
    notification_ids : List[int]
        Lista de IDs de notificaciones a eliminar
        Ejemplo: [1, 2, 3, 4, 5]
    
    db : Session
        Sesión de base de datos
    
    current_user : User
        Usuario autenticado
    
    Retorna:
    --------
    dict:
        Mensaje de confirmación con cantidad eliminada
    """
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to perform this action"
        )
    
    # Validar que la lista no esté vacía
    if not notification_ids or len(notification_ids) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La lista de notificaciones no puede estar vacía"
        )
    
    # Contar cuántas existen antes de eliminar
    existing_count = db.query(FormApprovalNotification).filter(
        FormApprovalNotification.id.in_(notification_ids)
    ).count()
    
    if existing_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No se encontraron notificaciones para eliminar"
        )
    
    # ⭐ ELIMINAR TODO DE UNA SOLA VEZ
    db.query(FormApprovalNotification).filter(
        FormApprovalNotification.id.in_(notification_ids)
    ).delete(synchronize_session=False)
    
    db.commit()
    
    return {
        "message": f"{existing_count} notificaciones eliminadas correctamente",
        "deleted_count": existing_count
    }

@router.delete("/forms/{form_id}")
def delete_form_endpoint(form_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Endpoint para eliminar un formulario.
    """
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create forms"
        )
    return delete_form(db, form_id)

@router.get("/forms/{form_id}/relations")
def get_form_relations_endpoint(form_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Endpoint para obtener las relaciones de un formulario sin eliminarlo.
    """
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to view form relations"
        )
    
    form = db.query(Form).filter(Form.id == form_id).first()
    if not form:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Formulario no encontrado."
        )
    
    relations_info = analyze_form_relations(db, form_id)
    
    return relations_info


@router.post("/create_form_close_config", response_model=FormCloseConfigOut)
def create_form_close_config(config: FormCloseConfigCreate, db: Session = Depends(get_db)):
    """
    Crea una nueva configuración de cierre para un formulario
    """
    
    # Validar si ya existe una configuración para el form_id
    existing = db.query(FormCloseConfig).filter(FormCloseConfig.form_id == config.form_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe una configuración para este formulario.")
    
    # 🆕 Validar que si se selecciona una acción, haya al menos un email
    if config.send_download_link and (not config.download_link_recipients or len(config.download_link_recipients) == 0):
        raise HTTPException(status_code=400, detail="Se requiere al menos un destinatario para enviar enlace de descarga.")
    
    if config.send_pdf_attachment and (not config.email_recipients or len(config.email_recipients) == 0):
        raise HTTPException(status_code=400, detail="Se requiere al menos un destinatario para enviar PDF como adjunto.")
    
    if config.generate_report and (not config.report_recipients or len(config.report_recipients) == 0):
        raise HTTPException(status_code=400, detail="Se requiere al menos un destinatario para generar reporte.")
    
    # 🆕 Convertir listas a JSON para almacenar
    config_dict = config.dict()
    config_dict['download_link_recipients'] = json.dumps(config.download_link_recipients or [])
    config_dict['email_recipients'] = json.dumps(config.email_recipients or [])
    config_dict['report_recipients'] = json.dumps(config.report_recipients or [])
    
    new_config = FormCloseConfig(**config_dict)
    db.add(new_config)
    db.commit()
    db.refresh(new_config)
    
    return new_config


@router.get("/form_close_config/{form_id}", response_model=FormCloseConfigOut)
def get_form_close_config(form_id: int, db: Session = Depends(get_db)):
    """
    Obtiene la configuración de cierre de un formulario específico
    """
    config = db.query(FormCloseConfig).filter(FormCloseConfig.form_id == form_id).first()
    
    if not config:
        raise HTTPException(status_code=404, detail="No se encontró configuración de cierre para este formulario")
    
    return config


@router.put("/form_close_config/{form_id}", response_model=FormCloseConfigOut)
def update_form_close_config(
    form_id: int,
    config: FormCloseConfigCreate,
    db: Session = Depends(get_db)
):
    """
    Actualiza la configuración de cierre de un formulario
    """
    existing_config = db.query(FormCloseConfig).filter(FormCloseConfig.form_id == form_id).first()
    
    if not existing_config:
        raise HTTPException(status_code=404, detail="No existe configuración de cierre para este formulario")
    
    # 🆕 Validaciones con múltiples emails
    if config.send_download_link and (not config.download_link_recipients or len(config.download_link_recipients) == 0):
        raise HTTPException(status_code=400, detail="Se requiere al menos un destinatario para enviar enlace de descarga.")
    
    if config.send_pdf_attachment and (not config.email_recipients or len(config.email_recipients) == 0):
        raise HTTPException(status_code=400, detail="Se requiere al menos un destinatario para enviar PDF como adjunto.")
    
    if config.generate_report and (not config.report_recipients or len(config.report_recipients) == 0):
        raise HTTPException(status_code=400, detail="Se requiere al menos un destinatario para generar reporte.")
    
    # 🆕 Convertir listas a JSON
    config_dict = config.dict()
    config_dict['download_link_recipients'] = json.dumps(config.download_link_recipients or [])
    config_dict['email_recipients'] = json.dumps(config.email_recipients or [])
    config_dict['report_recipients'] = json.dumps(config.report_recipients or [])
    
    # Actualizar campos
    for key, value in config_dict.items():
        setattr(existing_config, key, value)
    
    db.commit()
    db.refresh(existing_config)
    
    return existing_config


@router.delete("/form_close_config/{form_id}")
def delete_form_close_config(form_id: int, db: Session = Depends(get_db)):
    """
    Elimina la configuración de cierre de un formulario
    """
    config = db.query(FormCloseConfig).filter(FormCloseConfig.form_id == form_id).first()
    
    if not config:
        raise HTTPException(status_code=404, detail="No existe configuración de cierre para este formulario")
    
    db.delete(config)
    db.commit()
    
    return {"message": "Configuración de cierre eliminada exitosamente"}

UPLOAD_FOLDER = "logo"
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg"}

# Asegurar que la carpeta exista
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@router.post("/upload-logo/")
async def upload_logo(file: UploadFile = File(...),current_user: User = Depends(get_current_user)):
    # Validar extensión
    if current_user == None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to get options"
            )
    filename = file.filename
    extension = filename.split(".")[-1].lower()

    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido. Solo se permiten: png, jpg, jpeg.")

    # Elimina cualquier imagen anterior en la carpeta
    for f in os.listdir(UPLOAD_FOLDER):
        os.remove(os.path.join(UPLOAD_FOLDER, f))

    # Guardar la nueva imagen
    file_path = os.path.join(UPLOAD_FOLDER, f"logo.{extension}")
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    return JSONResponse(content={"message": "Imagen subida correctamente", "path": file_path})


@router.get("/get-logo/")
def get_logo(current_user: User = Depends(get_current_user)):
    if current_user == None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to get options"
            )
    if not os.path.exists(UPLOAD_FOLDER):
        raise HTTPException(status_code=404, detail="No se encontró la carpeta de logo.")

    archivos = os.listdir(UPLOAD_FOLDER)
    if not archivos:
        raise HTTPException(status_code=404, detail="No hay imagen guardada.")

    # Tomar el primer archivo (solo debe haber uno)
    imagen_path = os.path.join(UPLOAD_FOLDER, archivos[0])
    return FileResponse(imagen_path, media_type="image/*", filename=archivos[0])



# Endpoint principal para obtener el logo
@router.get("/public-logo/")
def get_public_logo():
    if not os.path.exists(UPLOAD_FOLDER):
        raise HTTPException(status_code=404, detail="No se encontró la carpeta de logo.")
    
    archivos = os.listdir(UPLOAD_FOLDER)
    if not archivos:
        raise HTTPException(status_code=404, detail="No hay imagen guardada.")
    
    imagen_path = os.path.join(UPLOAD_FOLDER, archivos[0])
    
    # Verificar que el archivo realmente existe
    if not os.path.isfile(imagen_path):
        raise HTTPException(status_code=404, detail="Archivo de logo no encontrado.")
    
    extension = archivos[0].split(".")[-1].lower()
    
    # Mapeo más completo de tipos MIME
    media_type_map = {
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "gif": "image/gif",
        "webp": "image/webp",
        "svg": "image/svg+xml"
    }
    
    media_type = media_type_map.get(extension, "image/*")
    
    # Headers para mejor manejo de cache y CORS
    headers = {
        "Cache-Control": "public, max-age=300",  # Cache por 5 minutos
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "GET, HEAD, OPTIONS",
        "Access-Control-Allow-Headers": "*"
    }
    
    return FileResponse(
        imagen_path, 
        media_type=media_type,
        headers=headers
    )

# Endpoint adicional para verificar existencia (opcional pero recomendado)
@router.get("/public-logo/exists")
def check_logo_exists():
    """Verifica si existe un logo sin descargar el archivo"""
    try:
        if not os.path.exists(UPLOAD_FOLDER):
            return {"exists": False, "reason": "Carpeta no existe"}
            
        archivos = os.listdir(UPLOAD_FOLDER)
        if not archivos:
            return {"exists": False, "reason": "No hay archivos"}
            
        imagen_path = os.path.join(UPLOAD_FOLDER, archivos[0])
        if not os.path.isfile(imagen_path):
            return {"exists": False, "reason": "Archivo no válido"}
            
        return {
            "exists": True, 
            "filename": archivos[0],
            "url": "/forms/public-logo/"
        }
    except Exception as e:
        return {"exists": False, "reason": f"Error: {str(e)}"}
    
# Soporte para HEAD requests
@router.head("/public-logo/")
def head_public_logo():
    """HEAD request para verificar existencia sin descargar"""
    if not os.path.exists(UPLOAD_FOLDER):
        raise HTTPException(status_code=404, detail="No se encontró la carpeta de logo.")
        
    archivos = os.listdir(UPLOAD_FOLDER)
    if not archivos:
        raise HTTPException(status_code=404, detail="No hay imagen guardada.")
        
    imagen_path = os.path.join(UPLOAD_FOLDER, archivos[0])
    if not os.path.isfile(imagen_path):
        raise HTTPException(status_code=404, detail="Archivo de logo no encontrado.")
    
    # SOLUCIÓN 1: Usar Response de FastAPI/Starlette (recomendado)
    from fastapi import Response
    return Response(status_code=200, headers={
        "Cache-Control": "public, max-age=300",
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "GET, HEAD, OPTIONS",
        "Access-Control-Allow-Headers": "*"
    })

@router.put("/update_form_category/{form_id}/category")
def update_form_category(
    form_id: int,
    category_data: UpdateFormCategory,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user is None:
        raise HTTPException(status_code=403, detail="No tienes permiso")

    form = db.query(Form).filter(Form.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Formulario no encontrado")

    old_category_id = form.id_category

    if category_data.id_category is not None:
        category = db.query(FormCategory).filter(FormCategory.id == category_data.id_category).first()
        if not category:
            raise HTTPException(status_code=404, detail="Categoría no encontrada")

    form.id_category = category_data.id_category
    db.commit()
    db.refresh(form)

    # ✅ SOLO sincronizar si el usuario lo aceptó
    if (form.id_category != old_category_id 
        and form.id_category is not None 
        and category_data.sync_approvers):
        sync_form_approvals_from_category(
            db=db,
            form_id=form.id,
            category_id=form.id_category,
            replace=True
        )

    return {
        "message": "Categoría actualizada correctamente",
        "form_id": form.id,
        "new_category_id": form.id_category,
        "approvers_synced": category_data.sync_approvers and form.id_category is not None
    }

# Endpoints para categorías de formularios
@router.post("/categories/", response_model=FormCategoryResponse, status_code=status.HTTP_201_CREATED)
def create_category_endpoint(
    category: FormCategoryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return create_form_category(db, category)

# Obtener árbol completo
@router.get("/categories/tree", response_model=List[FormCategoryTreeResponse])
def get_tree_endpoint(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return get_category_tree(db)

# Obtener categorías de un nivel
@router.get("/categories/by-parent", response_model=List[FormCategoryResponse])
def get_by_parent_endpoint(
    parent_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    categories = get_categories_by_parent(db, parent_id)
    
    # Enriquecer con contadores
    result = []
    for cat in categories:
        forms_count = db.query(func.count(Form.id)).filter(Form.id_category == cat.id).scalar()
        cat_dict = FormCategoryResponse.from_orm(cat).dict()
        cat_dict['forms_count'] = forms_count
        cat_dict['children_count'] = len(cat.children)
        result.append(cat_dict)
    
    return result

# Obtener una categoría específica
@router.get("/categories/{category_id}", response_model=FormCategoryResponse)
def get_category_endpoint(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    category = db.query(FormCategory).filter(FormCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")
    
    forms_count = db.query(func.count(Form.id)).filter(Form.id_category == category_id).scalar()
    response = FormCategoryResponse.from_orm(category)
    response.forms_count = forms_count
    response.children_count = len(category.children)
    
    return response

# Obtener ruta/breadcrumb de una categoría
@router.get("/categories/{category_id}/path", response_model=List[FormCategoryResponse])
def get_path_endpoint(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return get_category_path(db, category_id)

# Actualizar categoría
@router.put("/categories/{category_id}", response_model=FormCategoryResponse)
def update_category_endpoint(
    category_id: int,
    category_update: FormCategoryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return update_form_category_1(db, category_id, category_update)

@router.get("/users/form_by_user/search")
def search_user_forms(
    search: str,
    filter_type: str = Query("all", pattern="^(all|user|response_user)$"),
    page: int = 1,
    page_size: int = 30,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not search or search.strip() == "":
        raise HTTPException(status_code=400, detail="El término de búsqueda no puede estar vacío")
    if page_size > 100:
        page_size = 100

    # Llamar a la función de servicio
    result = search_forms_by_user(db, current_user.id, search.strip(), filter_type, page, page_size)
    return result


# Mover categoría
@router.patch("/categories/{category_id}/move", response_model=FormCategoryResponse)
def move_category_endpoint(
    category_id: int,
    move_data: FormCategoryMove,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return move_category(db, category_id, move_data)

# Eliminar categoría
@router.delete("/categories/{category_id}")
def delete_category_endpoint(
    category_id: int,
    force: bool = Query(False, description="Eliminar incluyendo subcategorías"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return delete_form_category(db, category_id, force)

# Obtener formularios de una categoría
@router.get("/categories/{category_id}/forms")
def get_forms_by_category_endpoint(
    category_id: int,
    page: int = 1,
    page_size: int = 30,
    include_subcategories: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if page_size > 100:
        page_size = 100
    
    offset = (page - 1) * page_size
    
    if not include_subcategories:
        base_query = db.query(Form).filter(Form.id_category == category_id)
    else:
        def get_all_descendant_ids(cat_id):
            ids = [cat_id]
            children = db.query(FormCategory).filter(FormCategory.parent_id == cat_id).all()
            for child in children:
                ids.extend(get_all_descendant_ids(child.id))
            return ids
        
        all_ids = get_all_descendant_ids(category_id)
        base_query = db.query(Form).filter(Form.id_category.in_(all_ids))
    
    total_count = base_query.count()
    forms = base_query.offset(offset).limit(page_size).all()
    total_pages = (total_count + page_size - 1) // page_size
    
    # Convertir a diccionarios manualmente
    items = []
    for form in forms:
        form_dict = {
            "id": form.id,
            "title": form.title,
            "format_type": form.format_type.value if hasattr(form.format_type, 'value') else form.format_type,
            "is_enabled": form.is_enabled,
            "user_id": form.user_id,
            "description": form.description,
            "created_at": form.created_at.isoformat() if form.created_at else None,
            "id_category": form.id_category,
            
        }
        items.append(form_dict)
    
    return {
        "items": items,
        "total": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "category_id": category_id,
        "include_subcategories": include_subcategories
    }

@router.get("/categories/{category_id}/user_forms")
def get_user_forms_by_category_endpoint(
    category_id: int,
    page: int = 1,
    page_size: int = 30,
    include_subcategories: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Retorna los formularios de una categoría que están ASIGNADOS al usuario autenticado.

    - **category_id**: ID de la categoría
    - **page**: Número de página (por defecto 1)
    - **page_size**: Cantidad de registros por página (por defecto 30, máximo 100)
    - **include_subcategories**: Incluir formularios de subcategorías (por defecto False)
    - **Requiere autenticación.**
    """
    if not current_user:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to access forms"
        )
    
    if page_size > 100:
        page_size = 100
    
    offset = (page - 1) * page_size
    
    # Determinar IDs de categorías a buscar
    if not include_subcategories:
        category_ids = [category_id]
    else:
        def get_all_descendant_ids(cat_id):
            ids = [cat_id]
            children = db.query(FormCategory).filter(FormCategory.parent_id == cat_id).all()
            for child in children:
                ids.extend(get_all_descendant_ids(child.id))
            return ids
        
        category_ids = get_all_descendant_ids(category_id)
    
    # Query base: Formularios asignados al usuario Y en la categoría
    base_query = (
        db.query(Form)
        .join(FormModerators, Form.id == FormModerators.form_id)
        .options(
            defer(Form.form_design),
            joinedload(Form.category)
        )
        .filter(FormModerators.user_id == current_user.id)  # ← Solo asignados al usuario
        .filter(Form.id_category.in_(category_ids))  # ← Y en la categoría
    )
    
    total_count = base_query.count()
    forms = base_query.offset(offset).limit(page_size).all()
    total_pages = (total_count + page_size - 1) // page_size
    
    # Convertir a diccionarios
    items = []
    for form in forms:
        form_dict = {
            "id": form.id,
            "title": form.title,
            "format_type": form.format_type.value if hasattr(form.format_type, 'value') else str(form.format_type),
            "is_enabled": form.is_enabled,
            "user_id": form.user_id,
            "description": form.description,
            "created_at": form.created_at.isoformat() if form.created_at else None,
            "id_category": form.id_category,
            "category": {
                "id": form.category.id,
                "parent_id": form.category.parent_id,
                "is_expanded": form.category.is_expanded,
                "color": form.category.color,
                "updated_at": form.category.updated_at.isoformat() if form.category.updated_at else None,
                "name": form.category.name,
                "description": form.category.description,
                "order": form.category.order,
                "icon": form.category.icon,
                "created_at": form.category.created_at.isoformat()
            } if form.category else None
        }
        items.append(form_dict)
    
    return {
        "items": items,
        "total": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "category_id": category_id,
        "include_subcategories": include_subcategories
    }
   
@router.get("/categories/{category_id}/user_completed_forms")
def get_user_completed_forms_by_category_endpoint(
    category_id: int,
    page: int = 1,
    page_size: int = 30,
    include_subcategories: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Retorna los formularios de una categoría que han sido COMPLETADOS/RESPONDIDOS por el usuario autenticado.

    - **category_id**: ID de la categoría
    - **page**: Número de página (por defecto 1)
    - **page_size**: Cantidad de registros por página (por defecto 30, máximo 100)
    - **include_subcategories**: Incluir formularios de subcategorías (por defecto False)
    - **Requiere autenticación.**
    """
    if not current_user:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to access forms"
        )
    
    if page_size > 100:
        page_size = 100
    
    offset = (page - 1) * page_size
    
    # Determinar IDs de categorías a buscar
    if not include_subcategories:
        category_ids = [category_id]
    else:
        def get_all_descendant_ids(cat_id):
            ids = [cat_id]
            children = db.query(FormCategory).filter(FormCategory.parent_id == cat_id).all()
            for child in children:
                ids.extend(get_all_descendant_ids(child.id))
            return ids
        
        category_ids = get_all_descendant_ids(category_id)
    
    # Query base: Formularios respondidos por el usuario Y en la categoría
    base_query = (
        db.query(Form)
        .join(Response, Form.id == Response.form_id)  # ← Join con Response en lugar de FormModerators
        .options(
            defer(Form.form_design),
            joinedload(Form.category)
        )
        .filter(Response.user_id == current_user.id)  # ← Solo respondidos por el usuario
        .filter(Form.id_category.in_(category_ids))  # ← Y en la categoría
        .distinct()  # ← Evitar duplicados si tiene múltiples respuestas al mismo form
    )
    
    total_count = base_query.count()
    forms = base_query.offset(offset).limit(page_size).all()
    total_pages = (total_count + page_size - 1) // page_size
    
    # Convertir a diccionarios (mismo formato que user_forms)
    items = []
    for form in forms:
        form_dict = {
            "id": form.id,
            "title": form.title,
            "format_type": form.format_type.value if hasattr(form.format_type, 'value') else str(form.format_type),
            "is_enabled": form.is_enabled,
            "user_id": form.user_id,
            "description": form.description,
            "created_at": form.created_at.isoformat() if form.created_at else None,
            "id_category": form.id_category,
            "category": {
                "id": form.category.id,
                "parent_id": form.category.parent_id,
                "is_expanded": form.category.is_expanded,
                "color": form.category.color,
                "updated_at": form.category.updated_at.isoformat() if form.category.updated_at else None,
                "name": form.category.name,
                "description": form.category.description,
                "order": form.category.order,
                "icon": form.category.icon,
                "created_at": form.category.created_at.isoformat()
            } if form.category else None
        }
        items.append(form_dict)
    
    return {
        "items": items,
        "total": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "category_id": category_id,
        "include_subcategories": include_subcategories
    }
    
@router.get("/api/forms/{form_id}/response/{response_id}/details")
async def get_response_details_json(
    form_id: int,
    response_id: int,
    db: Session = Depends(get_db)
):
    """
    Endpoint API que devuelve los detalles de una respuesta en formato JSON.
    Para ser consumido por el frontend Astro.
    No requiere autenticación.
    """
    try:
        # Obtener la respuesta específica
        stmt = (
            select(Response)
            .where(
                Response.id == response_id,
                Response.form_id == form_id
            )
            .options(
                joinedload(Response.answers).joinedload(Answer.question),
                joinedload(Response.approvals).joinedload(ResponseApproval.user),
                joinedload(Response.form),
                joinedload(Response.user)  # ✅ AGREGAR ESTA LÍNEA
            )
        )
        
        response = db.execute(stmt).unique().scalar_one_or_none()
        
        if not response:
            raise HTTPException(status_code=404, detail="Respuesta no encontrada")
        
        # Obtener el estado de aprobación
        approval_result = get_response_approval_status(response.approvals)
        
        # Procesar las aprobaciones
        processed_approvals = []
        for approval in response.approvals:
            user_info = approval.user
            processed_approvals.append({
                'approval_id': approval.id,
                'sequence_number': approval.sequence_number,
                'is_mandatory': approval.is_mandatory,
                'reconsideration_requested': approval.reconsideration_requested,
                'status': approval.status.value,
                'reviewed_at': approval.reviewed_at.isoformat() if approval.reviewed_at else None,
                'message': approval.message,
                'user': {
                    'name': user_info.name,
                    'email': user_info.email,
                    'nickname': user_info.nickname,
                    'num_document': user_info.num_document
                }
            })
        
        # Ordenar aprobaciones por sequence_number
        processed_approvals.sort(key=lambda x: x.get('sequence_number', 0))
        
        # Procesar respuestas del formulario
        form_answers = []
        for answer in response.answers:
            form_answers.append({
                'question_text': answer.question.question_text,
                'answer_text': answer.answer_text,
                'question_type': answer.question.question_type,
                'file_path': answer.file_path
            })
        
        # ✅ NUEVA SECCIÓN: Procesar información del usuario que respondió
        responded_by = None
        if response.user:
            responded_by = {
                'id': response.user.id,
                'name': response.user.name,
                'email': response.user.email,
                'nickname': response.user.nickname,
                'num_document': response.user.num_document
            }
        
        # Preparar respuesta JSON
        response_data = {
            'response_id': response.id,
            'submitted_at': response.submitted_at.isoformat() if response.submitted_at else None,
            'approval_status': approval_result["status"],
            'message': approval_result["message"],
            'responded_by': responded_by,  # ✅ AGREGAR ESTA LÍNEA
            'form': {
                'id': response.form.id,
                'title': response.form.title,
                'description': response.form.description
            },
            'approvals': processed_approvals,
            'answers': form_answers
        }
        
        return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error al obtener detalles de respuesta: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail="Error interno del servidor"
        )


@router.put("/forms/{form_id}/basic-info")
def update_form_basic_info(
    form_id: int,
    form_data: UpdateFormBasicInfo,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Actualiza el título, descripción y/o tipo de formato de un formulario.
    Solo el propietario del formulario puede realizar esta acción.
    """
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permiso para actualizar este formulario"
        )

    # Buscar el formulario
    form = db.query(Form).filter(Form.id == form_id).first()
    if not form:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Formulario no encontrado"
        )

    # Validar que al menos un campo esté presente
    if form_data.title is None and form_data.description is None and form_data.format_type is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Debe proporcionar al menos un campo para actualizar"
        )

    # Actualizar los campos proporcionados
    if form_data.title is not None:
        form.title = form_data.title
    
    if form_data.description is not None:
        form.description = form_data.description
    
    if form_data.format_type is not None:
        form.format_type = form_data.format_type

    db.commit()
    db.refresh(form)

    return {
        "message": "Información del formulario actualizada correctamente",
        "form_id": form.id,
        "title": form.title,
        "description": form.description,
        "format_type": form.format_type.value
    }


@router.patch("/forms/{form_id}/status")
def update_form_status(
    form_id: int,
    status_update: FormStatusUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Endpoint para habilitar o deshabilitar un formulario.
    Solo usuarios con rol ADMIN pueden usar este endpoint.
    """
    # Validar que el usuario sea administrador
    if current_user.user_type.name != UserType.admin.name:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only administrators can enable or disable forms"
        )
    
    return toggle_form_status(db, form_id, status_update.is_enabled)

@router.get("/by-question/{question_id}", response_model=List[FormResponseBitacora])
def get_forms_by_question(question_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Devuelve todos los formularios que contienen la pregunta con el ID especificado.
    """
    if str(current_user.user_type.value) not in ["admin", "creator"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permisos para acceder a las palabras clave."
        )

    # Buscar la pregunta
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(status_code=404, detail="Pregunta no encontrada.")

    # Acceder a los formularios relacionados (gracias a la relación many-to-many)
    forms = question.forms

    if not forms:
        raise HTTPException(status_code=404, detail="Esta pregunta no pertenece a ningún formulario.")

    return forms


INSTRUCTIVOS_FOLDER = "./form_instructivos"
os.makedirs(INSTRUCTIVOS_FOLDER, exist_ok=True)

# ==================== SUBIR INSTRUCTIVO ====================
@router.put("/{form_id}/upload-instructivos")
async def upload_form_instructivos(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    files: List[UploadFile] = File(..., description="Archivos de instructivos (cualquier tipo)"),
    descriptions: str = FastAPIForm(..., description="JSON array con descripciones para cada archivo")
):
    """
    Sube múltiples archivos de instructivos para un formulario.
    Acepta cualquier tipo de archivo (PDF, DOC, imágenes, videos, etc.) sin límite de tamaño.
    
    - **form_id**: ID del formulario
    - **files**: Lista de archivos
    - **descriptions**: JSON string con array de descripciones (una por archivo)
    
    Ejemplo de descriptions:
    ["Guía de llenado en PDF", "Video tutorial", "Ejemplo en imagen"]
    """
    try:
        # Verificar que el usuario está autenticado
        if current_user is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="User not authenticated"
            )

        # 1. Buscar el formulario
        form = db.query(Form).filter(Form.id == form_id).first()
        
        if not form:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Form with id {form_id} not found"
            )
        
        # 2. Verificar que el usuario es el dueño del formulario
        if form.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't have permission to modify this form"
            )

        # 3. Parsear las descripciones
        try:
            descriptions_list = json.loads(descriptions)
            if not isinstance(descriptions_list, list):
                raise ValueError("Descriptions must be a JSON array")
        except json.JSONDecodeError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid JSON format for descriptions"
            )

        # 4. Validar que hay la misma cantidad de archivos y descripciones
        if len(files) != len(descriptions_list):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Number of files ({len(files)}) must match number of descriptions ({len(descriptions_list)})"
            )

        # 5. Obtener la lista existente de instructivos (si hay)
        existing_instructivos = []
        if form.instructivo_url:
            try:
                # Si ya existe, intentar parsear como JSON
                if isinstance(form.instructivo_url, str):
                    existing_instructivos = json.loads(form.instructivo_url)
                elif isinstance(form.instructivo_url, list):
                    existing_instructivos = form.instructivo_url
            except:
                # Si falla, empezar con lista vacía
                existing_instructivos = []

        # 6. Procesar cada archivo
        new_instructivos = []
        uploaded_files_info = []

        for idx, (file, description) in enumerate(zip(files, descriptions_list)):
            # Validar que el archivo no esté vacío
            if not file.filename:
                continue

            try:
                # Generar nombre único para el archivo
                file_extension = os.path.splitext(file.filename)[1]
                unique_filename = f"instructivo_{form_id}_{uuid.uuid4()}{file_extension}"
                file_path = os.path.join(INSTRUCTIVOS_FOLDER, unique_filename)
                
                # Guardar el archivo
                content = await file.read()
                with open(file_path, "wb") as f:
                    f.write(content)
                
                # Crear objeto de instructivo
                instructivo_obj = {
                    "url": file_path,
                    "description": description.strip() if description else "Sin descripción",
                    "original_name": file.filename,
                    "file_type": file.content_type or "unknown",
                    "size": len(content)
                }
                
                new_instructivos.append(instructivo_obj)
                uploaded_files_info.append({
                    "name": file.filename,
                    "size": len(content),
                    "type": file.content_type
                })
                
                print(f"✅ Archivo {idx + 1} guardado: {file_path}")
                
            except Exception as e:
                # Si falla la subida de un archivo, continuar con los demás
                print(f"⚠️ Error al subir archivo {file.filename}: {str(e)}")
                continue

        # 7. Combinar instructivos existentes con los nuevos
        all_instructivos = existing_instructivos + new_instructivos

        # 8. Guardar en la base de datos como JSON
        form.instructivo_url = json.dumps(all_instructivos, ensure_ascii=False)
        db.commit()
        db.refresh(form)

        return {
            "message": f"{len(new_instructivos)} instructivo(s) uploaded successfully",
            "form_id": form_id,
            "uploaded_count": len(new_instructivos),
            "total_instructivos": len(all_instructivos),
            "uploaded_files": uploaded_files_info,
            "all_instructivos": all_instructivos
        }

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Unexpected error: {str(e)}"
        )

# ==================== OBTENER INSTRUCTIVOS ====================

@router.get("/{form_id}/instructivos")
async def get_form_instructivos(
    form_id: int,
    db: Session = Depends(get_db)
):
    """
    Obtiene la lista de instructivos de un formulario.
    """
    try:
        form = db.query(Form).filter(Form.id == form_id).first()
        
        if not form:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Form with id {form_id} not found"
            )

        instructivos = []
        if form.instructivo_url:
            try:
                if isinstance(form.instructivo_url, str):
                    instructivos = json.loads(form.instructivo_url)
                elif isinstance(form.instructivo_url, list):
                    instructivos = form.instructivo_url
            except:
                instructivos = []

        return {
            "form_id": form_id,
            "instructivos": instructivos,
            "count": len(instructivos)
        }

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Unexpected error: {str(e)}"
        )


# ==================== ENDPOINT 2: GUARDAR MENSAJE DE ALERTA ====================

@router.put("/{form_id}/update-alert-message")
async def update_form_alert_message(
    form_id: int,
    request: AlertMessageRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Actualiza el mensaje de alerta de un formulario.
    
    - **form_id**: ID del formulario
    - **alert_message**: Texto del mensaje de alerta
    """
    try:
        # Verificar que el usuario está autenticado
        if current_user is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="User not authenticated"
            )

        # 1. Buscar el formulario
        form = db.query(Form).filter(Form.id == form_id).first()
        
        if not form:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Form with id {form_id} not found"
            )
        
        # 2. Verificar que el usuario es el dueño del formulario
        if form.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't have permission to modify this form"
            )

        # 3. Validar que el mensaje no esté vacío (opcional, puedes quitarlo si quieres permitir vacío)
        if not request.alert_message or request.alert_message.strip() == "":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Alert message cannot be empty"
            )

        # 4. Actualizar el mensaje de alerta
        form.alert_message = request.alert_message.strip()
        db.commit()
        db.refresh(form)

        print(f"✅ Mensaje de alerta actualizado para el form {form_id}")

        return {
            "message": "Alert message updated successfully",
            "form_id": form_id,
            "alert_message": form.alert_message
        }

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Unexpected error: {str(e)}"
        )

@router.get("/files/download-instructivo")
async def download_instructivo(
    file_path: str = Query(..., description="Ruta del archivo a descargar"),
    current_user: User = Depends(get_current_user)
):
    try:
        if current_user is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="User not authenticated"
            )

        # 🔥 SOLO ESTO: Normalizar la ruta
        file_path = file_path.replace('\\', '/')
        
        if not os.path.exists(file_path):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found"
            )

        filename = os.path.basename(file_path)
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/octet-stream',
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error: {str(e)}"
        )
        
@router.get("/{form_id}/alert-message")
async def get_form_alert_message(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtiene el mensaje de alerta de un formulario.
    
    - **form_id**: ID del formulario
    """
    try:
        # Verificar que el usuario está autenticado
        if current_user is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="User not authenticated"
            )

        # Buscar el formulario
        form = db.query(Form).filter(Form.id == form_id).first()
        
        if not form:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Form with id {form_id} not found"
            )

        # Retornar el mensaje de alerta (puede ser None o vacío)
        return {
            "form_id": form_id,
            "alert_message": form.alert_message,
            "has_alert": bool(form.alert_message and form.alert_message.strip())
        }

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Unexpected error: {str(e)}"
        )
        
# ==================== OBTENER UN FORMULARIO (para ver detalles) ====================

@router.get("/get_form_details/{form_id}")
async def get_form_details(
    form_id: int,
    db: Session = Depends(get_db)
):
    """
    Obtiene los detalles de un formulario específico.
    Incluyendo: título, descripción, instructivos, alert_message, etc.
    """
    try:
        form = db.query(Form).filter(Form.id == form_id).first()
        
        if not form:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Form with id {form_id} not found"
            )

        # Parsear instructivos si existen
        instructivos = []
        if form.instructivo_url:
            try:
                if isinstance(form.instructivo_url, str):
                    instructivos = json.loads(form.instructivo_url)
                elif isinstance(form.instructivo_url, list):
                    instructivos = form.instructivo_url
            except:
                instructivos = []

        return {
            "id": form.id,
            "title": form.title,
            "description": form.description,
            "alert_message": form.alert_message or "",
            "instructivos": instructivos,
            "created_at": form.created_at,
            "user_id": form.user_id,
            "is_enabled": form.is_enabled,
            "format_type": form.format_type
        }

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Unexpected error: {str(e)}"
        )


# ==================== ELIMINAR UN INSTRUCTIVO ESPECÍFICO ====================

@router.delete("/{form_id}/instructivos/{instructivo_index}")
async def delete_instructivo(
    form_id: int,
    instructivo_index: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Elimina un instructivo específico de un formulario.
    
    - **form_id**: ID del formulario
    - **instructivo_index**: Índice del instructivo a eliminar (0, 1, 2, etc.)
    """
    try:
        # Verificar que el usuario está autenticado
        if current_user is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="User not authenticated"
            )

        # 1. Buscar el formulario
        form = db.query(Form).filter(Form.id == form_id).first()
        
        if not form:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Form with id {form_id} not found"
            )
        
        # 2. Verificar que el usuario es el dueño
        if form.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't have permission to modify this form"
            )

        # 3. Obtener instructivos actuales
        instructivos = []
        if form.instructivo_url:
            try:
                if isinstance(form.instructivo_url, str):
                    instructivos = json.loads(form.instructivo_url)
                elif isinstance(form.instructivo_url, list):
                    instructivos = form.instructivo_url
            except:
                instructivos = []

        # 4. Validar índice
        if instructivo_index < 0 or instructivo_index >= len(instructivos):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Instructivo index {instructivo_index} out of range"
            )

        # 5. Obtener el archivo a eliminar
        instructivo_to_delete = instructivos[instructivo_index]
        file_path = instructivo_to_delete.get("url")

        # 6. Eliminar el archivo del servidor (si existe)
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
                print(f"✅ Archivo eliminado del servidor: {file_path}")
            except Exception as e:
                print(f"⚠️ Error al eliminar archivo: {str(e)}")
                # No interrumpimos si falla la eliminación del archivo

        # 7. Eliminar de la lista
        instructivos.pop(instructivo_index)

        # 8. Guardar cambios
        form.instructivo_url = json.dumps(instructivos, ensure_ascii=False)
        db.commit()
        db.refresh(form)

        print(f"✅ Instructivo {instructivo_index} eliminado del form {form_id}")

        return {
            "message": "Instructivo eliminado correctamente",
            "form_id": form_id,
            "instructivos_restantes": instructivos,
            "total_instructivos": len(instructivos)
        }

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Unexpected error: {str(e)}"
        )


# ==================== ELIMINAR MENSAJE DE ALERTA ====================

@router.delete("/{form_id}/alert-message")
async def delete_alert_message(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Elimina el mensaje de alerta de un formulario.
    
    - **form_id**: ID del formulario
    """
    try:
        # Verificar que el usuario está autenticado
        if current_user is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="User not authenticated"
            )

        # 1. Buscar el formulario
        form = db.query(Form).filter(Form.id == form_id).first()
        
        if not form:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Form with id {form_id} not found"
            )
        
        # 2. Verificar que el usuario es el dueño
        if form.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't have permission to modify this form"
            )

        # 3. Verificar que existe un mensaje
        if not form.alert_message:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No alert message to delete"
            )

        # 4. Eliminar el mensaje
        form.alert_message = None
        db.commit()
        db.refresh(form)

        print(f"✅ Mensaje de alerta eliminado del form {form_id}")

        return {
            "message": "Alert message deleted successfully",
            "form_id": form_id,
            "alert_message": ""
        }

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Unexpected error: {str(e)}"
        )
        
@router.post(
    "/movimientos",
    response_model=FormMovimientoResponse,
    status_code=status.HTTP_201_CREATED
)
def create_form_movimiento_endpoint(
    movimiento: FormMovimientoBase,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Crear un nuevo movimiento.

    Solo usuarios con rol `creator` o `admin` pueden crear movimientos.
    """
    if current_user.user_type.name not in [UserType.creator.name, UserType.admin.name]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to create movimientos"
        )

    return create_form_movimiento(
        db=db,
        movimiento=movimiento,
        user_id=current_user.id
    )


@router.get("/movimientos/all", response_model=List[dict])
def get_form_movimientos_endpoint(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Retorna una lista básica de movimientos de formularios.

    - id
    - title
    - description
    """

    # 🔐 Validación de roles permitidos
    if current_user.user_type.name not in [
        UserType.creator.name,
        UserType.admin.name
    ]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to view movimientos"
        )

    movimientos = get_all_form_movimientos_basic(db)

    if not movimientos:
        return []  # 👈 mejor que 404 para listas

    return movimientos

def get_question_labels_from_form_design(form_design: list) -> dict:
    """
    Retorna un dict:
    {
        question_id: label
    }
    """
    labels = {}

    if not form_design:
        return labels

    for element in form_design:
        question_id = element.get("id_question")
        props = element.get("props", {})

        if question_id and "label" in props:
            labels[question_id] = props["label"]

    return labels
@router.get(
    "/movimientos/{movement_id}/answers",
    status_code=status.HTTP_200_OK
)
def get_answers_by_movement(
    movement_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 🔐 Validar roles
    if current_user.user_type.name not in [
        UserType.creator.name,
        UserType.admin.name
    ]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not have permission to view movement answers"
        )

    movimiento = db.query(FormMovimientos).filter(
        FormMovimientos.id == movement_id,
        FormMovimientos.is_enabled == True
    ).first()

    if not movimiento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Movimiento no encontrado"
        )

    if not movimiento.form_ids or not movimiento.question_ids:
        return {
            "movement_id": movimiento.id,
            "title": movimiento.title,
            "forms": []
        }

    forms = db.query(Form).filter(
        Form.id.in_(movimiento.form_ids)
    ).all()

    result = []

    for form in forms:
        # 🔑 Mapear labels desde el form_design
        question_labels = get_question_labels_from_form_design(
            form.form_design or []
        )

        responses = db.query(Response).filter(
            Response.form_id == form.id,
            Response.status == ResponseStatus.submitted
        ).all()

        form_responses = []

        for response in responses:
            answers = db.query(Answer).join(Question).filter(
                Answer.response_id == response.id,
                Answer.question_id.in_(movimiento.question_ids)
            ).all()

            if not answers:
                continue

            form_responses.append({
                "response_id": response.id,
                "submitted_at": response.submitted_at,
                "answers": [
                    {
                        "question_id": a.question.id,
                        "question_text": a.question.question_text,
                        "question_label": question_labels.get(
                            a.question.id,
                            a.question.question_text
                        ),
                        "alias": {
                            "id": a.question.alias.id,
                            "name": a.question.alias.name,
                            "description": a.question.alias.description
                        } if a.question.alias else None,  # 👈 NUEVO
                        "answer_text": a.answer_text,
                        "file_path": a.file_path
                    }
                    for a in answers
                ]
            })

        if form_responses:
            result.append({
                "form_id": form.id,
                "form_title": form.title,
                "responses": form_responses
            })

    return {
        "movement_id": movimiento.id,
        "title": movimiento.title,
        "description": movimiento.description,
        "forms": result
    }

@router.delete("/movimientos/{movement_id}", status_code=status.HTTP_200_OK)
def delete_movement(
    movement_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    movement = (
        db.query(FormMovimientos)
        .filter(
            FormMovimientos.id == movement_id,
            FormMovimientos.user_id == current_user.id
        )
        .first()
    )

    if not movement:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Movimiento no encontrado"
        )

    db.delete(movement)
    db.commit()

    return {
        "message": "Movimiento eliminado correctamente"
    }
    

@router.post("/responses/related-last-answer")
def get_related_last_answers(
    payload: RelatedAnswerRequest,
    db: Session = Depends(get_db)
):
    # 1️⃣ Obtener response_id donde la pregunta MATCH tenga el valor dado
    response_ids = (
        db.query(Answer.response_id)
        .join(Response, Response.id == Answer.response_id)
        .filter(
            Response.form_id == payload.form_id,
            Answer.question_id == payload.question_id_match,
            Answer.answer_text == payload.value_base
        )
        .distinct()
        .all()
    )

    response_ids = [r.response_id for r in response_ids]

    if not response_ids:
        return []

    # 2️⃣ Buscar relación de la pregunta lookup
    relation = (
        db.query(QuestionTableRelation)
        .filter(
            QuestionTableRelation.question_id == payload.question_id_lookup
        )
        .first()
    )

    if not relation or not relation.related_question_id:
        raise HTTPException(
            status_code=404,
            detail="La pregunta no tiene relación definida en QuestionTableRelation"
        )

    related_question_id = relation.related_question_id

    # 3️⃣ Obtener TODAS las últimas respuestas en UNA SOLA QUERY (optimizado)
    # Subquery para obtener el máximo ID de Answer por cada response_id
    max_answer_subquery = (
        db.query(
            Answer.response_id,
            func.max(Answer.id).label('max_id')
        )
        .filter(
            Answer.response_id.in_(response_ids),
            Answer.question_id == related_question_id
        )
        .group_by(Answer.response_id)
        .subquery()
    )

    # Query principal que obtiene las respuestas usando la subquery
    last_answers = (
        db.query(Answer)
        .join(
            max_answer_subquery,
            Answer.id == max_answer_subquery.c.max_id
        )
        .all()
    )

    # Construir resultados
    results = [
        {
            "response_id": answer.response_id,
            "question_id": related_question_id,
            "answer_id": answer.id,
            "answer_text": answer.answer_text,
            "file_path": answer.file_path
        }
        for answer in last_answers
    ]

    return results

@router.post("/send-answers-by-email")
def send_answers_by_email(payload: SendResponseEmailRequest):

    ok = send_response_answers_email(
        to_emails=payload.email_to,
        form_title=payload.form_title,
        response_id=payload.response_id,
        answers=payload.answers
    )

    if not ok:
        raise HTTPException(
            status_code=500,
            detail="No se pudo enviar el correo"
        )

    return {
        "status": "ok",
        "sent_to": payload.email_to
    }

